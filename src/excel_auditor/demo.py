"""Generate realistic demo workbooks.

`financial_model_v1.xlsx` is a small but structurally realistic financial
model. `financial_model_v2.xlsx` is the same model with ten deliberate
anomalies (documented inline below); the test-suite asserts the engine
detects each of them.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.workbook.defined_name import DefinedName

MONTHS = [
    "Jan", "Feb", "Mar", "Apr", "May", "Jun",
    "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
]

BOLD = Font(bold=True)


def build_demo_workbook(path: Path, *, with_anomalies: bool) -> Path:
    wb = Workbook()

    # ------------------------------------------------------------ Assumptions
    ws = wb.active
    ws.title = "Assumptions"
    ws["A1"] = "Assumptions"
    ws["A1"].font = BOLD
    ws["A2"] = "Unit price"
    ws["B2"] = 25.0
    ws["B2"].number_format = "#,##0.00"
    ws["A3"] = "Monthly growth"
    # Anomaly 6: a changed assumption value.
    ws["B3"] = 0.07 if with_anomalies else 0.05
    ws["B3"].number_format = "0.0%"
    ws["A4"] = "COGS % of revenue"
    ws["B4"] = 0.4
    ws["B4"].number_format = "0.0%"
    ws["A5"] = "Monthly opex"
    ws["B5"] = 15000
    ws["A6"] = "Tax rate"
    ws["B6"] = 0.1
    ws["B6"].number_format = "0.0%"
    ws["A7"] = "Starting cash"
    ws["B7"] = 50000
    # Hidden row containing data (both versions - part of the base model).
    ws["A9"] = "Internal: previous scenario, do not delete"
    ws.row_dimensions[9].hidden = True
    # Hidden column containing data (both versions).
    ws["E2"] = "reviewer note: check growth vs 2024 actuals"
    ws.column_dimensions["E"].hidden = True

    # ------------------------------------------------------ Revenue Forecast
    rf = wb.create_sheet("Revenue Forecast")
    for col, header in zip("ABCD", ["Month", "Units", "Price", "Revenue"], strict=True):
        rf[f"{col}1"] = header
        rf[f"{col}1"].font = BOLD
    for i, month in enumerate(MONTHS):
        row = i + 2
        rf[f"A{row}"] = month
        rf[f"B{row}"] = 100 + 4 * i
        rf[f"C{row}"] = "=Assumptions!$B$2"
        # Anomaly 1: formula overwritten with a hardcoded number (row 7 = June... May).
        if with_anomalies and row == 7:
            rf[f"D{row}"] = 2600
        else:
            rf[f"D{row}"] = f"=B{row}*C{row}"
    rf["A14"] = "Total"
    rf["A14"].font = BOLD
    # Anomaly 3: total range excludes the final month.
    rf["D14"] = "=SUM(D2:D12)" if with_anomalies else "=SUM(D2:D13)"

    # ------------------------------------------------------------------- P&L
    pl = wb.create_sheet("P&L")
    pl["A1"] = "P&L Statement"
    pl["A1"].font = BOLD
    labels = {
        2: "Revenue",
        3: "COGS",
        4: "Gross Profit",
        5: "Opex",
        6: "EBITDA",
        7: "Tax",
        8: "Net Profit",
    }
    for row, label in labels.items():
        pl[f"A{row}"] = label
    pl["B2"] = "='Revenue Forecast'!D14"
    # Anomaly 10: formula change with broad downstream impact (also hardcodes a rate).
    pl["B3"] = "=B2*0.45" if with_anomalies else "=B2*Assumptions!$B$4"
    pl["B4"] = "=B2-B3"
    pl["B5"] = "=Assumptions!$B$5*12"
    pl["B6"] = "=B4-B5"
    pl["B7"] = "=MAX(0,B6*Assumptions!$B$6)"
    pl["B8"] = "=B6-B7"
    for row in range(2, 9):
        pl[f"B{row}"].number_format = "#,##0"
    # Anomaly 7: formatting-only change (bold + different number format, same formula).
    if with_anomalies:
        pl["B6"].font = BOLD
        pl["B6"].number_format = "#,##0.00"

    # -------------------------------------------------------------- Cash Flow
    cf = wb.create_sheet("Cash Flow")
    for col, header in zip("ABC", ["Month", "Net Cash", "Cumulative Cash"], strict=True):
        cf[f"{col}1"] = header
        cf[f"{col}1"].font = BOLD
    for i in range(12):
        row = i + 2
        cf[f"A{row}"] = i + 1
        # Anomaly 2: one copied formula references the wrong row (D8 instead of D9).
        if with_anomalies and row == 9:
            cf[f"B{row}"] = "='Revenue Forecast'!D8*0.9-Assumptions!$B$5"
        else:
            cf[f"B{row}"] = f"='Revenue Forecast'!D{row}*0.9-Assumptions!$B$5"
        if row == 2:
            cf[f"C{row}"] = "=Assumptions!$B$7+B2"
        else:
            cf[f"C{row}"] = f"=C{row - 1}+B{row}"
    # Anomaly 8: a broken reference.
    if with_anomalies:
        cf["D1"] = "Adjustment"
        cf["D2"] = "=#REF!+B2"

    # ---------------------------------------------------------------- Summary
    summary = wb.create_sheet("Summary")
    summary["A1"] = "Company Summary"
    summary["A1"].font = BOLD
    summary.merge_cells("A1:B1")
    summary["A3"] = "Total Revenue"
    summary["B3"] = "='Revenue Forecast'!D14"
    summary["A4"] = "EBITDA"
    summary["B4"] = "='P&L'!B6"
    summary["A5"] = "Net Profit"
    summary["B5"] = "='P&L'!B8"
    summary["A6"] = "Ending Cash"
    summary["B6"] = "='Cash Flow'!C13"
    if with_anomalies:
        # Anomaly 5: reference to an external workbook.
        summary["A7"] = "Benchmark Revenue"
        summary["B7"] = "='[Benchmarks.xlsx]Data'!B2"
        # Anomaly 9: a volatile function.
        summary["A8"] = "Report date"
        summary["B8"] = "=TODAY()"

    # ----------------------------------------------------------------- Фактури
    # Anomaly 11: cent-level rounding drift. The line items carry sub-cent
    # precision, so their displayed values add to 4,456.00 while the true sum
    # 4,456.005 displays as 4,456.01 — and the total was hand-typed to match.
    if with_anomalies:
        inv = wb.create_sheet("Фактури")
        for col, header in zip("AB", ["Услуга", "Сума"], strict=True):
            inv[f"{col}1"] = header
            inv[f"{col}1"].font = BOLD
        items = [
            ("Консултации май", 1234.564),
            ("Абонамент поддръжка", 2345.333),
            ("Лицензи (преизчислени)", 876.108),
        ]
        for row, (label, amount) in enumerate(items, start=2):
            inv[f"A{row}"] = label
            inv[f"B{row}"] = amount
            inv[f"B{row}"].number_format = '#,##0.00 "лв"'
        inv["A5"] = "Общо"
        inv["A5"].font = BOLD
        inv["B5"] = 4456.01  # hand-corrected to the rounded true sum
        inv["B5"].number_format = '#,##0.00 "лв"'

    # Anomaly 4: a hidden sheet containing data.
    if with_anomalies:
        adj = wb.create_sheet("Adjustments")
        adj["A1"] = "Manual adjustment"
        adj["B1"] = 5000
        adj["A2"] = "Agreed with CFO 2025-11-14"
        adj.sheet_state = "hidden"

    # Named range (both versions).
    defined = DefinedName("TaxRate", attr_text="Assumptions!$B$6")
    try:
        wb.defined_names["TaxRate"] = defined
    except TypeError:  # pragma: no cover - older openpyxl API
        wb.defined_names.add(defined)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def generate_demo_workbooks(directory: Path | str) -> tuple[Path, Path]:
    directory = Path(directory)
    v1 = build_demo_workbook(directory / "financial_model_v1.xlsx", with_anomalies=False)
    v2 = build_demo_workbook(directory / "financial_model_v2.xlsx", with_anomalies=True)
    return v1, v2
