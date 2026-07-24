"""Repeated-range inconsistency detection.

Scans every column (and row) for runs of cells sharing the same normalized
formula. A single-cell interruption between two runs of the same pattern is
flagged: a hardcoded constant (formula overwritten), a structurally different
formula, or a blank (missing formula).

Thresholds are deliberately conservative: only single-cell gaps with at least
three pattern cells around them are reported, which keeps false positives low.
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import Any, Literal

from openpyxl.utils import get_column_letter

from ..models import CellRecord, SheetInventory, WorkbookInventory
from ..parsing.formula_normalizer import normalize_formula

AnomalyKind = Literal["overwritten_constant", "inconsistent_formula", "missing_formula"]

MIN_PATTERN_CELLS = 3  # total pattern cells surrounding the gap
MIN_LONGEST_RUN = 2  # at least one side must have >= 2 consecutive pattern cells


@dataclass(frozen=True)
class PatternAnomaly:
    kind: AnomalyKind
    sheet_name: str
    coordinate: str
    expected_normalized: str
    orientation: Literal["column", "row"]
    run_before: int
    run_after: int
    actual_formula: str | None
    actual_value: Any
    shifted_by: int | None  # offset that would make the formula match the pattern


@dataclass(frozen=True)
class _Run:
    normalized: str
    start: int  # position along the scan axis (row number or column number)
    end: int

    @property
    def length(self) -> int:
        return self.end - self.start + 1


def _collect_runs(cells_by_pos: dict[int, CellRecord]) -> list[_Run]:
    runs: list[_Run] = []
    current_norm: str | None = None
    current_start = 0
    previous_pos: int | None = None

    for pos in sorted(cells_by_pos):
        record = cells_by_pos[pos]
        norm = record.normalized_formula if record.is_formula else None
        contiguous = previous_pos is not None and pos == previous_pos + 1
        if norm is not None and norm == current_norm and contiguous:
            runs[-1] = _Run(norm, current_start, pos)
        elif norm is not None:
            runs.append(_Run(norm, pos, pos))
            current_norm = norm
            current_start = pos
        else:
            current_norm = None
        previous_pos = pos
    return runs


def _shift_that_matches(
    record: CellRecord, expected: str, orientation: str
) -> int | None:
    """If normalizing the formula from a shifted anchor reproduces the expected
    pattern, the formula was probably copied from / points at the wrong row/col."""
    if record.formula is None:
        return None
    for delta in (-3, -2, -1, 1, 2, 3):
        if orientation == "column":
            shifted = normalize_formula(
                record.formula, row=record.row + delta, column=record.column
            )
        else:
            shifted = normalize_formula(
                record.formula, row=record.row, column=record.column + delta
            )
        if shifted == expected:
            return delta
    return None


def _classify_gap(
    gap_record: CellRecord | None,
    run_before: _Run,
    run_after: _Run,
    sheet_name: str,
    coordinate: str,
    orientation: Literal["column", "row"],
) -> PatternAnomaly | None:
    expected = run_before.normalized

    def _anomaly(
        kind: AnomalyKind,
        actual_formula: str | None,
        actual_value: Any,
        shifted_by: int | None,
    ) -> PatternAnomaly:
        return PatternAnomaly(
            kind=kind,
            sheet_name=sheet_name,
            coordinate=coordinate,
            expected_normalized=expected,
            orientation=orientation,
            run_before=run_before.length,
            run_after=run_after.length,
            actual_formula=actual_formula,
            actual_value=actual_value,
            shifted_by=shifted_by,
        )

    if gap_record is None:
        return _anomaly("missing_formula", None, None, None)
    if gap_record.is_numeric_constant:
        return _anomaly("overwritten_constant", None, gap_record.value, None)
    if gap_record.is_formula and gap_record.normalized_formula != expected:
        return _anomaly(
            "inconsistent_formula",
            gap_record.formula,
            gap_record.value,
            _shift_that_matches(gap_record, expected, orientation),
        )
    return None  # text/bool gaps are ignored (likely intentional labels)


def _scan_line(
    sheet: SheetInventory,
    cells_by_pos: dict[int, CellRecord],
    orientation: Literal["column", "row"],
    fixed_index: int,
    populated_positions: set[int],
) -> list[PatternAnomaly]:
    anomalies: list[PatternAnomaly] = []
    runs = _collect_runs(cells_by_pos)
    # Pair consecutive runs OF THE SAME pattern: a divergent formula in the gap
    # forms its own run, so plain adjacent pairing would miss it.
    runs_by_norm: dict[str, list[_Run]] = {}
    for run in runs:
        runs_by_norm.setdefault(run.normalized, []).append(run)
    for norm_runs in runs_by_norm.values():
        for before, after in zip(norm_runs, norm_runs[1:], strict=False):
            gap_size = after.start - before.end - 1
            if gap_size != 1:
                continue
            if before.length + after.length < MIN_PATTERN_CELLS:
                continue
            if max(before.length, after.length) < MIN_LONGEST_RUN:
                continue
            gap_pos = before.end + 1
            if orientation == "column":
                coordinate = f"{get_column_letter(fixed_index)}{gap_pos}"
            else:
                coordinate = f"{get_column_letter(gap_pos)}{fixed_index}"
            gap_record = sheet.cells.get(coordinate)
            # Spacer suppression: a blank gap in an entirely blank row/column
            # is layout, not a missing formula.
            if gap_record is None and gap_pos not in populated_positions:
                continue
            anomaly = _classify_gap(
                gap_record, before, after, sheet.name, coordinate, orientation
            )
            if anomaly is not None:
                anomalies.append(anomaly)
    return anomalies


def detect_pattern_anomalies(inventory: WorkbookInventory) -> list[PatternAnomaly]:
    results: dict[tuple[str, str], PatternAnomaly] = {}

    for sheet in inventory.sheets:
        by_column: dict[int, dict[int, CellRecord]] = {}
        by_row: dict[int, dict[int, CellRecord]] = {}
        for record in sheet.cells.values():
            by_column.setdefault(record.column, {})[record.row] = record
            by_row.setdefault(record.row, {})[record.column] = record
        populated_rows = set(by_row)
        populated_columns = set(by_column)

        for column, cells in by_column.items():
            for anomaly in _scan_line(sheet, cells, "column", column, populated_rows):
                results.setdefault((sheet.name, anomaly.coordinate), anomaly)
        for row, cells in by_row.items():
            for anomaly in _scan_line(sheet, cells, "row", row, populated_columns):
                results.setdefault((sheet.name, anomaly.coordinate), anomaly)

    return list(results.values())
