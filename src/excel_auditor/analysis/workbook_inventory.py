"""Build a typed WorkbookInventory from a loaded openpyxl workbook."""

from __future__ import annotations

import hashlib
import logging
from pathlib import Path

from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.worksheet import Worksheet

from ..config import Settings
from ..models import (
    CellRecord,
    NamedRange,
    SheetInventory,
    SheetVisibility,
    TableInfo,
    WorkbookInventory,
)
from ..parsing.formula_normalizer import normalize_formula
from ..parsing.workbook_loader import LoadedWorkbook, load_workbook_safe

logger = logging.getLogger(__name__)


def _formula_text(value: object) -> str | None:
    if isinstance(value, str) and value.startswith("="):
        return value
    if isinstance(value, ArrayFormula):
        text = value.text or ""
        return text if text.startswith("=") else "=" + text
    return None


def _style_signature(cell) -> str:
    """Compact signature of visually meaningful formatting for diffing."""
    font = cell.font
    fill = cell.fill
    fg = ""
    try:
        if fill is not None and fill.fgColor is not None and fill.fgColor.rgb:
            fg = str(fill.fgColor.rgb)
    except (AttributeError, TypeError):
        fg = ""
    return "|".join(
        [
            str(cell.number_format or ""),
            f"b{int(bool(font and font.b))}",
            f"i{int(bool(font and font.i))}",
            f"u{(font.u or '') if font else ''}",
            fg,
        ]
    )


def _sheet_visibility(state: str | None) -> SheetVisibility:
    try:
        return SheetVisibility(state or "visible")
    except ValueError:
        return SheetVisibility.VISIBLE


def _hidden_rows(ws: Worksheet) -> list[int]:
    return sorted(int(idx) for idx, dim in ws.row_dimensions.items() if dim.hidden)


def _hidden_columns(ws: Worksheet) -> list[str]:
    from openpyxl.utils import column_index_from_string, get_column_letter

    letters: set[str] = set()
    for key, dim in ws.column_dimensions.items():
        if not dim.hidden:
            continue
        lo = dim.min or column_index_from_string(key)
        hi = dim.max or column_index_from_string(key)
        for idx in range(lo, hi + 1):
            letters.add(get_column_letter(idx))
    return sorted(letters)


def _build_sheet(
    ws_formulas: Worksheet, ws_values: Worksheet | None, index: int
) -> SheetInventory:
    cells: dict[str, CellRecord] = {}
    for row in ws_formulas.iter_rows():
        for cell in row:
            raw_value = cell.value
            formula = _formula_text(raw_value)
            if raw_value is None and formula is None:
                continue

            cached = None
            if ws_values is not None:
                try:
                    cached = ws_values[cell.coordinate].value
                except (KeyError, IndexError, ValueError):
                    cached = None

            normalized = None
            if formula is not None:
                normalized = normalize_formula(formula, row=cell.row, column=cell.column)

            cells[cell.coordinate] = CellRecord(
                coordinate=cell.coordinate,
                row=cell.row,
                column=cell.column,
                data_type=str(cell.data_type),
                value=cached if formula is not None else raw_value,
                formula=formula,
                normalized_formula=normalized,
                number_format=str(cell.number_format or ""),
                style_signature=_style_signature(cell),
            )

    tables: list[TableInfo] = []
    try:
        # NB: openpyxl's TableList.items() yields (name, ref-string); plain
        # dict access is needed to reach the Table object and its metadata.
        table_map = ws_formulas.tables or {}
        for name in table_map:
            table = dict.get(table_map, name)
            if table is None:
                continue
            ref = table.ref if hasattr(table, "ref") else str(table)
            # Exact structure from the Table XML; None → 0 per decision D14
            # (openpyxl defaults: headerRowCount=1, totalsRowCount=None).
            header_count = getattr(table, "headerRowCount", 1)
            totals_count = getattr(table, "totalsRowCount", None)
            tables.append(
                TableInfo(
                    name=str(name),
                    ref=str(ref),
                    sheet_name=ws_formulas.title,
                    header_row_count=int(header_count) if header_count is not None else 0,
                    totals_row_count=int(totals_count) if totals_count is not None else 0,
                )
            )
    except (AttributeError, TypeError):
        pass

    return SheetInventory(
        name=ws_formulas.title,
        index=index,
        visibility=_sheet_visibility(ws_formulas.sheet_state),
        max_row=ws_formulas.max_row or 0,
        max_column=ws_formulas.max_column or 0,
        dimensions=ws_formulas.calculate_dimension() if cells else None,
        protected=bool(ws_formulas.protection and ws_formulas.protection.sheet),
        merged_ranges=sorted(str(r) for r in ws_formulas.merged_cells.ranges),
        hidden_rows=_hidden_rows(ws_formulas),
        hidden_columns=_hidden_columns(ws_formulas),
        tables=tables,
        cells=cells,
    )


def _named_ranges(wb) -> list[NamedRange]:
    result: list[NamedRange] = []

    def _collect(container, scope: str | None) -> None:
        try:
            items = list(container.items())
        except (AttributeError, TypeError):
            items = [(dn.name, dn) for dn in getattr(container, "definedName", [])]
        for name, dn in items:
            refers = getattr(dn, "attr_text", None) or getattr(dn, "value", None)
            result.append(
                NamedRange(
                    name=str(name),
                    refers_to=str(refers) if refers is not None else None,
                    scope=scope,
                    hidden=bool(getattr(dn, "hidden", False)),
                )
            )

    if getattr(wb, "defined_names", None) is not None:
        _collect(wb.defined_names, None)
    for ws in wb.worksheets:
        sheet_names = getattr(ws, "defined_names", None)
        if sheet_names:
            _collect(sheet_names, ws.title)
    return result


def _workbook_link_targets(wb) -> list[str]:
    targets: list[str] = []
    for link in getattr(wb, "_external_links", None) or []:
        target = getattr(getattr(link, "file_link", None), "Target", None)
        if target:
            targets.append(str(target))
    return targets


def _external_links(wb, sheets: list[SheetInventory], zip_external_parts: int) -> list[str]:
    """One consistent list of external dependencies.

    Combines workbook-level link targets with targets derived from formula
    text, so the overview count always matches what the external-dependency
    rule reports. Numeric markers like `[1]Sheet1!A1` are resolved against the
    workbook-level link list (Excel's `[n]` is an index into it).
    """
    workbook_targets = _workbook_link_targets(wb)
    combined: list[str] = list(workbook_targets)

    from ..parsing.formula_tokenizer import reference_tokens
    from ..parsing.reference_parser import parse_reference

    for sheet in sheets:
        for record in sheet.formula_cells:
            if record.formula is None:
                continue
            for token in reference_tokens(record.formula):
                parsed = parse_reference(token.value)
                if parsed is None or not parsed.is_external:
                    continue
                marker = parsed.external or ""
                if marker.isdigit() and 0 < int(marker) <= len(workbook_targets):
                    target = workbook_targets[int(marker) - 1]
                else:
                    target = marker
                if target and target not in combined:
                    combined.append(target)

    if not combined and zip_external_parts:
        combined = [f"<{zip_external_parts} external link part(s) present>"]
    return combined


def _file_sha256(path: Path) -> str:
    """Full SHA-256 hex digest of the workbook file bytes (deterministic id)."""
    digest = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            digest.update(chunk)
    return digest.hexdigest()


def build_inventory(
    loaded: LoadedWorkbook,
    *,
    workbook_id: str | None = None,
    filename: str | None = None,
) -> WorkbookInventory:
    wb = loaded.formulas
    wbv = loaded.values

    sheets: list[SheetInventory] = []
    for index, ws in enumerate(wb.worksheets):
        ws_values = None
        try:
            ws_values = wbv[ws.title]
        except KeyError:
            logger.warning("Sheet missing in values pass: %r", ws.title)
        sheets.append(_build_sheet(ws, ws_values, index))

    calc_mode = None
    full_precision = None
    calc = getattr(wb, "calculation", None)
    if calc is not None:
        calc_mode = getattr(calc, "calcMode", None)
        raw_precision = getattr(calc, "fullPrecision", None)
        if raw_precision is not None:
            full_precision = bool(raw_precision)

    security = getattr(wb, "security", None)
    protected = bool(
        security
        and (getattr(security, "lockStructure", False) or getattr(security, "lockWindows", False))
    )

    return WorkbookInventory(
        workbook_id=workbook_id or _file_sha256(loaded.path),
        filename=filename or loaded.path.name,
        file_size=loaded.file_size,
        sheets=sheets,
        named_ranges=_named_ranges(wb),
        external_links=_external_links(wb, sheets, loaded.zip_facts.external_link_parts),
        has_macros=loaded.zip_facts.has_vba,
        has_data_connections=loaded.zip_facts.has_data_connections,
        calculation_mode=str(calc_mode) if calc_mode else None,
        full_precision=full_precision,
        workbook_protected=protected,
    )


def inventory_from_path(
    path: Path,
    *,
    settings: Settings | None = None,
    workbook_id: str | None = None,
    filename: str | None = None,
) -> WorkbookInventory:
    """Convenience: validate, load and inventory a workbook file."""
    loaded = load_workbook_safe(Path(path), settings)
    return build_inventory(loaded, workbook_id=workbook_id, filename=filename)
