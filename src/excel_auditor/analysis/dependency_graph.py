"""Cell dependency graph.

Nodes are (sheet_name, coordinate) pairs. Edges run from a referenced cell to
the formula cell that depends on it. External workbook references are kept in
a side table and never traversed. Traversal is BFS with a visited set, so
cycles cannot crash it.
"""

from __future__ import annotations

import logging
import re
from collections import defaultdict, deque
from collections.abc import Iterator
from dataclasses import dataclass, field

from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import coordinate_from_string

from ..models import DependencyImpact, NamedRange, SheetInventory, TableInfo, WorkbookInventory
from ..parsing.formula_tokenizer import reference_tokens
from ..parsing.reference_parser import (
    CellRef,
    ParsedReference,
    _split_sheet_prefix,
    parse_reference,
)

logger = logging.getLogger(__name__)

Key = tuple[str, str]  # (sheet name, coordinate) e.g. ("P&L", "B3")

# Heuristics for "output-like" cells (see plan: no perfect semantic classification).
OUTPUT_SHEET_KEYWORDS = (
    "summary",
    "dashboard",
    "report",
    "p&l",
    "pnl",
    "balance",
    "cash flow",
    "cashflow",
    "forecast",
    "budget",
)
OUTPUT_LABEL_KEYWORDS = (
    "total",
    "ebitda",
    "revenue",
    "profit",
    "cash",
    "margin",
    "net",
    "subtotal",
)
MANY_DEPENDENTS_THRESHOLD = 15


@dataclass
class DependencyGraph:
    dependents: dict[Key, set[Key]] = field(default_factory=lambda: defaultdict(set))
    precedents: dict[Key, set[Key]] = field(default_factory=lambda: defaultdict(set))
    external_refs: dict[Key, set[str]] = field(default_factory=lambda: defaultdict(set))
    # Cells whose formulas reference themselves (directly or via a range that
    # covers them). Kept out of dependents/precedents so BFS still terminates.
    self_loops: set[Key] = field(default_factory=set)
    # Formula cells containing name-like tokens (defined names, structured
    # table references) that could NOT be resolved into edges: unknown names,
    # #REF!/formula-valued names, unsupported table item specifiers. Cells
    # whose name tokens all resolved — or were understood constants — are not
    # recorded. While non-empty, dependent counts computed from this graph may
    # be understated for ANY cell — the queried cell could be the unresolved
    # name's target.
    unresolved_name_cells: set[Key] = field(default_factory=set)
    truncated_ranges: int = 0

    # ------------------------------------------------------------------ build

    @classmethod
    def build(
        cls, inventory: WorkbookInventory, *, max_range_cells: int = 10_000
    ) -> DependencyGraph:
        graph = cls()
        sheet_by_upper = {s.name.upper(): s for s in inventory.sheets}
        # Name/table lookups; first definition wins on duplicates (inventory
        # order is deterministic).
        global_names: dict[str, NamedRange] = {}
        scoped_names: dict[tuple[str, str], NamedRange] = {}
        for named in inventory.named_ranges:
            if named.scope is None:
                global_names.setdefault(named.name.upper(), named)
            else:
                scoped_names.setdefault((named.scope.upper(), named.name.upper()), named)
        tables_by_upper: dict[str, TableInfo] = {}
        for sheet in inventory.sheets:
            for table in sheet.tables:
                tables_by_upper.setdefault(table.name.upper(), table)

        for sheet in inventory.sheets:
            for record in sheet.formula_cells:
                dependent: Key = (sheet.name, record.coordinate)
                assert record.formula is not None
                for token in reference_tokens(record.formula):
                    parsed = parse_reference(token.value)
                    if parsed is not None:
                        graph._add_reference_edges(
                            dependent, sheet, parsed, sheet_by_upper, max_range_cells
                        )
                        continue
                    # Defined name or structured table reference.
                    areas = _resolve_name_token(
                        token.value,
                        sheet,
                        sheet_by_upper,
                        global_names,
                        scoped_names,
                        tables_by_upper,
                    )
                    if areas is None:
                        # Genuinely unresolvable: impacts computed from this
                        # graph may be understated.
                        graph.unresolved_name_cells.add(dependent)
                        continue
                    for area in areas:
                        graph._add_reference_edges(
                            dependent, sheet, area, sheet_by_upper, max_range_cells
                        )
        return graph

    def _add_reference_edges(
        self,
        dependent: Key,
        formula_sheet: SheetInventory,
        parsed: ParsedReference,
        sheet_by_upper: dict[str, SheetInventory],
        max_range_cells: int,
    ) -> None:
        """Add edges for one parsed reference (direct or name/table-resolved)."""
        if parsed.is_external:
            self.external_refs[dependent].add(parsed.raw)
            return

        resolved = (
            sheet_by_upper.get(parsed.sheet.upper()) if parsed.sheet is not None else formula_sheet
        )
        if resolved is None:
            return  # reference to a sheet we cannot resolve
        target_sheet = resolved

        r1 = parsed.start.row
        c1 = parsed.start.column
        r2 = parsed.end.row if parsed.end else r1
        c2 = parsed.end.column if parsed.end else c1
        # Whole-column/row refs: clamp to the sheet's used range.
        r1 = 1 if r1 is None else r1
        c1 = 1 if c1 is None else c1
        r2 = target_sheet.max_row if r2 is None else r2
        c2 = target_sheet.max_column if c2 is None else c2
        r1, r2 = min(r1, r2), max(r1, r2)
        c1, c2 = min(c1, c2), max(c1, c2)
        r2 = min(r2, max(target_sheet.max_row, r1))
        c2 = min(c2, max(target_sheet.max_column, c1))

        area = (r2 - r1 + 1) * (c2 - c1 + 1)
        if area > max_range_cells:
            self.truncated_ranges += 1
            r2 = min(r2, r1 + max_range_cells - 1)
            c2 = c1  # keep it a bounded strip rather than exploding

        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                referenced: Key = (target_sheet.name, f"{get_column_letter(c)}{r}")
                if referenced == dependent:
                    self.self_loops.add(dependent)
                    continue
                self.dependents[referenced].add(dependent)
                self.precedents[dependent].add(referenced)

    # -------------------------------------------------------------- traversal

    def transitive_dependents(self, key: Key) -> set[Key]:
        seen: set[Key] = set()
        queue: deque[Key] = deque(self.dependents.get(key, ()))
        while queue:
            node = queue.popleft()
            if node in seen:
                continue
            seen.add(node)
            queue.extend(self.dependents.get(node, ()))
        return seen

    def cycles(self) -> list[list[Key]]:
        """All strongly connected components of size > 1 (plus self-loops).

        Iterative Tarjan - no recursion, safe on large graphs.
        """
        adjacency = self.dependents
        index: dict[Key, int] = {}
        low: dict[Key, int] = {}
        on_stack: set[Key] = set()
        stack: list[Key] = []
        counter = 0
        components: list[list[Key]] = []

        for root in list(adjacency.keys()):
            if root in index:
                continue
            work: list[tuple[Key, Iterator[Key]]] = [(root, iter(adjacency.get(root, ())))]
            index[root] = low[root] = counter
            counter += 1
            stack.append(root)
            on_stack.add(root)

            while work:
                node, children = work[-1]
                advanced = False
                for child in children:
                    if child not in index:
                        index[child] = low[child] = counter
                        counter += 1
                        stack.append(child)
                        on_stack.add(child)
                        work.append((child, iter(adjacency.get(child, ()))))
                        advanced = True
                        break
                    if child in on_stack:
                        low[node] = min(low[node], index[child])
                if advanced:
                    continue
                work.pop()
                if work:
                    parent = work[-1][0]
                    low[parent] = min(low[parent], low[node])
                if low[node] == index[node]:
                    component: list[Key] = []
                    while True:
                        w = stack.pop()
                        on_stack.discard(w)
                        component.append(w)
                        if w == node:
                            break
                    if len(component) > 1:
                        components.append(component)

        # Self-edges are kept out of the adjacency (so traversal terminates),
        # so self-loops surface here as single-cell components instead.
        in_multi_cell_component = {node for component in components for node in component}
        for key in sorted(self.self_loops):
            if key not in in_multi_cell_component:
                components.append([key])
        return components


# ------------------------------------------------- name/table resolution

_TABLE_TOKEN_RE = re.compile(
    r"^(?P<table>[A-Za-z_\\][A-Za-z0-9_.\\]*)\[(?P<column>.*)\]$", re.DOTALL
)
# Literal-valued defined names (`="0.05"`, `=0.05`, `=TRUE`). Fully understood:
# no cell target exists, so uses produce no edges and no unknown marker.
_CONSTANT_NAME_RE = re.compile(
    r'^(?:"[^"]*"|[-+]?\d+(?:\.\d+)?(?:[eE][-+]?\d+)?|TRUE|FALSE)$', re.IGNORECASE
)


def _split_top_level_areas(text: str) -> list[str]:
    """Split a multi-area refers_to on commas outside quoted sheet names/strings."""
    areas: list[str] = []
    buf: list[str] = []
    in_single = in_double = False
    for ch in text:
        if ch == "'" and not in_double:
            in_single = not in_single
        elif ch == '"' and not in_single:
            in_double = not in_double
        if ch == "," and not in_single and not in_double:
            areas.append("".join(buf))
            buf = []
        else:
            buf.append(ch)
    areas.append("".join(buf))
    return areas


def _resolve_defined_name(
    named: NamedRange, sheet_by_upper: dict[str, SheetInventory]
) -> list[ParsedReference] | None:
    """refers_to -> parsed areas; [] for constants; None when unresolvable.

    All-or-nothing: if any area of a multi-area refers_to cannot be resolved
    soundly, the whole name stays unresolved (marker) — partial edges would
    understate impact while looking resolved.
    """
    if named.refers_to is None:
        return None
    text = named.refers_to.strip()
    if text.startswith("="):
        text = text[1:].strip()
    if not text:
        return None
    if "#REF!" in text.upper():
        return None  # broken name: its impact genuinely is unknown
    if _CONSTANT_NAME_RE.match(text):
        return []
    parsed_areas: list[ParsedReference] = []
    for area in _split_top_level_areas(text):
        parsed = parse_reference(area.strip())
        if parsed is None:
            return None  # e.g. formula-valued name (=OFFSET(...))
        if not parsed.is_external:
            # refers_to areas are sheet-qualified; an unknown or missing sheet
            # would silently contribute no edges, so treat it as unresolved.
            if parsed.sheet is None or parsed.sheet.upper() not in sheet_by_upper:
                return None
        parsed_areas.append(parsed)
    return parsed_areas or None


def _resolve_table_reference(
    table: TableInfo,
    column_spec: str,
    sheet_by_upper: dict[str, SheetInventory],
) -> list[ParsedReference] | None:
    """`Table1[Col]` / `Table1[]` -> the data-row sub-range; None when unsure.

    The inventory records only the table's full ref, so we assume Excel's
    default single header row and detect a totals row by its self-referential
    aggregate formula (e.g. =SUBTOTAL(109,Table1[Col])) in the ref's last row.
    Item specifiers (#All/#Headers/..., @, multi-column spans, escaped column
    names) are not resolved and keep the unknown marker.
    """
    spec = column_spec.strip()
    if any(ch in spec for ch in "[]#@:'"):
        return None
    sheet = sheet_by_upper.get(table.sheet_name.upper())
    if sheet is None:
        return None
    ref = parse_reference(table.ref)
    if ref is None or ref.is_external or ref.sheet is not None:
        return None  # table refs are plain local ranges like "A1:C5"
    r1, c1 = ref.start.row, ref.start.column
    r2 = ref.end.row if ref.end else r1
    c2 = ref.end.column if ref.end else c1
    if r1 is None or c1 is None or r2 is None or c2 is None:
        return None
    data_r1 = r1 + 1  # header row excluded (Excel's default headerRowCount=1)
    data_r2 = r2
    table_upper = table.name.upper()
    if data_r2 >= data_r1:
        for c in range(c1, c2 + 1):
            record = sheet.cells.get(f"{get_column_letter(c)}{data_r2}")
            if (
                record is not None
                and record.formula is not None
                and table_upper in record.formula.upper()
            ):
                data_r2 -= 1  # totals row detected: exclude it
                break
    if data_r2 < data_r1:
        return []  # header-only table: no data cells, nothing understated
    if spec == "":
        # Table1[] = the whole data region.
        return [
            ParsedReference(
                raw=table.ref,
                sheet=table.sheet_name,
                external=None,
                start=CellRef(row=data_r1, column=c1),
                end=CellRef(row=data_r2, column=c2),
            )
        ]
    wanted = spec.casefold()
    matches = [
        c
        for c in range(c1, c2 + 1)
        if (header := sheet.cells.get(f"{get_column_letter(c)}{r1}")) is not None
        and isinstance(header.value, str)
        and header.value.strip().casefold() == wanted
    ]
    if len(matches) != 1:
        return None  # unknown or duplicated column header: ambiguous
    return [
        ParsedReference(
            raw=table.ref,
            sheet=table.sheet_name,
            external=None,
            start=CellRef(row=data_r1, column=matches[0]),
            end=CellRef(row=data_r2, column=matches[0]),
        )
    ]


def _resolve_name_token(
    token: str,
    formula_sheet: SheetInventory,
    sheet_by_upper: dict[str, SheetInventory],
    global_names: dict[str, NamedRange],
    scoped_names: dict[tuple[str, str], NamedRange],
    tables_by_upper: dict[str, TableInfo],
) -> list[ParsedReference] | None:
    """Resolve a non-A1 reference token to concrete areas.

    Returns a list of areas to add edges for ([] when the token is fully
    understood but targets no cells), or None — the caller then records the
    unknown-impact marker. Mis-resolution is worse than no resolution, so
    every uncertain path returns None.
    """
    token = token.strip()
    m = _TABLE_TOKEN_RE.match(token)
    if m is not None:
        table = tables_by_upper.get(m.group("table").upper())
        if table is None:
            return None
        return _resolve_table_reference(table, m.group("column"), sheet_by_upper)
    prefix, name_part = _split_sheet_prefix(token)
    key = name_part.strip().upper()
    if not key:
        return None
    if prefix is not None:
        if "[" in prefix:
            return None  # external-workbook name
        named = scoped_names.get((prefix.upper(), key))
    else:
        # A sheet-scoped name shadows a same-named workbook-scoped name for
        # formulas on its own sheet.
        named = scoped_names.get((formula_sheet.name.upper(), key)) or global_names.get(key)
    if named is None:
        return None
    return _resolve_defined_name(named, sheet_by_upper)


# ------------------------------------------------------------------- impact


def _looks_like_output(
    inventory: WorkbookInventory, graph: DependencyGraph, key: Key
) -> str | None:
    """Return a reason string when the cell looks like a model output, else None."""
    sheet_name, coordinate = key
    lowered = sheet_name.lower()
    for keyword in OUTPUT_SHEET_KEYWORDS:
        if keyword in lowered:
            return f"on output-like sheet '{sheet_name}'"

    sheet = inventory.sheet(sheet_name)
    if sheet is not None:
        try:
            col_letter, row = coordinate_from_string(coordinate)
            col = column_index_from_string(col_letter)
        except ValueError:
            return None
        neighbours: list[str] = []
        for delta in range(1, 4):
            if col - delta >= 1:
                neighbours.append(f"{get_column_letter(col - delta)}{row}")
            if row - delta >= 1:
                neighbours.append(f"{col_letter}{row - delta}")
        for coord in neighbours:
            record = sheet.cells.get(coord)
            if record is not None and isinstance(record.value, str):
                text = record.value.lower()
                if any(k in text for k in OUTPUT_LABEL_KEYWORDS):
                    return f"near label '{record.value}'"

    if len(graph.dependents.get(key, ())) >= MANY_DEPENDENTS_THRESHOLD:
        return "has many downstream dependents"
    return None


def impact_for(
    graph: DependencyGraph,
    inventory: WorkbookInventory,
    key: Key,
    *,
    max_output_scan: int = 5_000,
) -> DependencyImpact:
    direct = graph.dependents.get(key, set())
    transitive = graph.transitive_dependents(key)

    outputs: list[str] = []
    for candidate in sorted(transitive)[:max_output_scan]:
        reason = _looks_like_output(inventory, graph, candidate)
        if reason is not None:
            outputs.append(f"{candidate[0]}!{candidate[1]} ({reason})")
            if len(outputs) >= 10:
                break

    return DependencyImpact(
        direct_dependent_count=len(direct),
        transitive_dependent_count=len(transitive),
        affected_sheets=sorted({s for s, _ in transitive}),
        touches_outputs=bool(outputs),
        sample_output_cells=outputs,
        sample_direct_dependents=[f"{s}!{c}" for s, c in sorted(direct)[:20]],
        is_circular=key in transitive or key in graph.self_loops,
        has_unresolved_names=bool(graph.unresolved_name_cells),
    )
