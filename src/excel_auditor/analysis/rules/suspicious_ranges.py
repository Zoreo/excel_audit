"""Aggregation ranges that appear to exclude adjacent populated cells, and
formulas referencing blank cells."""

from __future__ import annotations

from openpyxl.utils import get_column_letter

from ...models import Confidence, Finding, Severity
from ...parsing.formula_tokenizer import function_names, reference_tokens
from ...parsing.reference_parser import parse_reference
from .base import AuditContext, Rule, register

AGGREGATION_FUNCTIONS = {
    "SUM",
    "AVERAGE",
    "MIN",
    "MAX",
    "COUNT",
    "COUNTA",
    "SUBTOTAL",
    "MEDIAN",
}
_MAX_EXAMPLES = 25


@register
class RangeExcludesAdjacentCellRule(Rule):
    rule_id = "EA-RNG-001"
    title = "Total range may exclude an adjacent populated cell"
    description = (
        "An aggregation range stops right before a populated cell - typically a "
        "total that was not extended after adding the latest row."
    )

    def run(self, ctx: AuditContext) -> list[Finding]:
        findings: list[Finding] = []
        for sheet in ctx.inventory.sheets:
            for record in sheet.formula_cells:
                assert record.formula is not None
                if not function_names(record.formula) & AGGREGATION_FUNCTIONS:
                    continue
                for token in reference_tokens(record.formula):
                    parsed = parse_reference(token.value)
                    if (
                        parsed is None
                        or parsed.is_external
                        or parsed.end is None
                        or parsed.start.row is None
                        or parsed.end.row is None
                        or parsed.start.column is None
                        or parsed.end.column is None
                    ):
                        continue
                    # Only same-sheet references can be checked against this sheet.
                    if parsed.sheet is not None and parsed.sheet.upper() != sheet.name.upper():
                        continue

                    vertical = parsed.start.column == parsed.end.column
                    horizontal = parsed.start.row == parsed.end.row
                    if vertical and not horizontal:
                        beyond_row = max(parsed.start.row, parsed.end.row) + 1
                        coordinate = f"{get_column_letter(parsed.start.column)}{beyond_row}"
                    elif horizontal and not vertical:
                        beyond_col = max(parsed.start.column, parsed.end.column) + 1
                        coordinate = f"{get_column_letter(beyond_col)}{parsed.start.row}"
                    else:
                        continue

                    if coordinate == record.coordinate:
                        continue  # the total cell itself sits right below/right - normal
                    beyond = sheet.cells.get(coordinate)
                    if beyond is None:
                        continue
                    if not (beyond.is_numeric_constant or beyond.is_formula):
                        continue
                    findings.append(
                        Finding(
                            rule_id=self.rule_id,
                            title=self.title,
                            description=(
                                f"'{sheet.name}'!{record.coordinate} aggregates "
                                f"{token.value}, but adjacent cell {coordinate} is populated "
                                "and not included."
                            ),
                            severity=Severity.HIGH,
                            confidence=Confidence.MEDIUM,
                            location=ctx.location(sheet.name, record.coordinate),
                            related_locations=[ctx.location(sheet.name, coordinate)],
                            evidence={
                                "formula": record.formula,
                                "range": token.value,
                                "excluded_cell": coordinate,
                            },
                            suggested_action=(
                                "Check whether the range should include the adjacent cell."
                            ),
                        )
                    )
                    if len(findings) >= _MAX_EXAMPLES:
                        return findings
        return findings


@register
class ReferencesBlankCellsRule(Rule):
    rule_id = "EA-RNG-002"
    title = "Formulas reference blank cells"
    description = "Direct references to empty cells may indicate deleted inputs."

    def run(self, ctx: AuditContext) -> list[Finding]:
        findings: list[Finding] = []
        for sheet in ctx.inventory.sheets:
            examples: list[dict[str, str]] = []
            for record in sheet.formula_cells:
                assert record.formula is not None
                for token in reference_tokens(record.formula):
                    parsed = parse_reference(token.value)
                    if (
                        parsed is None
                        or parsed.is_external
                        or parsed.is_range
                        or parsed.start.row is None
                        or parsed.start.column is None
                    ):
                        continue
                    target_sheet = sheet
                    if parsed.sheet is not None:
                        resolved = ctx.inventory.sheet(parsed.sheet)
                        if resolved is None:
                            continue
                        target_sheet = resolved
                    coordinate = (
                        f"{get_column_letter(parsed.start.column)}{parsed.start.row}"
                    )
                    # Only inside the used range: trailing blank space is normal.
                    if (
                        parsed.start.row > target_sheet.max_row
                        or parsed.start.column > target_sheet.max_column
                    ):
                        continue
                    if coordinate in target_sheet.cells:
                        continue
                    examples.append(
                        {
                            "cell": f"{sheet.name}!{record.coordinate}",
                            "references_blank": f"{target_sheet.name}!{coordinate}",
                        }
                    )
                    break  # one blank ref per formula cell is enough evidence
            if examples:
                findings.append(
                    Finding(
                        rule_id=self.rule_id,
                        title=self.title,
                        description=(
                            f"{len(examples)} formula(s) on '{sheet.name}' reference blank "
                            "cells inside the used range."
                        ),
                        severity=Severity.LOW,
                        confidence=Confidence.LOW,
                        location=ctx.location(sheet.name),
                        evidence={"examples": examples[:10]},
                        suggested_action=(
                            "Verify these blanks are intentional (e.g. optional inputs)."
                        ),
                    )
                )
        return findings
