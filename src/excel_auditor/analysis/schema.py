"""Schema detection: find tables, headers, column types and total rows.

Heuristic by design - the goal is a schema good enough for deterministic
querying plus explicit notes where the structure is ambiguous, never a claim
of perfect understanding.
"""

from __future__ import annotations

from datetime import date, datetime
from typing import Any

from openpyxl.utils import get_column_letter

from ..models import CellRecord, SheetInventory, WorkbookInventory
from ..models.schema import (
    ColumnSchema,
    ColumnType,
    SheetSchemaInfo,
    TableSchema,
    WorkbookSchema,
)
from .resolution import concepts_of, normalize

_MIN_TABLE_ROWS = 2
_HEADER_TEXT_SHARE = 0.6
_SAMPLE_LIMIT = 3

_TOTAL_KEYWORDS = {
    "total", "subtotal", "grand total", "sum",
    "общо", "итого", "всичко", "сума", "общ",
}
_BOOL_STRINGS = {"да", "не", "yes", "no", "true", "false", "y", "n"}
_CURRENCY_MARKERS = [
    ("€", "EUR"), ("eur", "EUR"),
    ("лв", "BGN"), ("bgn", "BGN"),
    ("$", "USD"), ("usd", "USD"),
    ("£", "GBP"), ("gbp", "GBP"),
]


def _is_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def _is_date(value: Any) -> bool:
    return isinstance(value, (datetime, date))


def _is_texty(value: Any) -> bool:
    return isinstance(value, str) and value.strip() != ""


def _currency_from_format(number_format: str | None) -> str | None:
    if not number_format:
        return None
    lowered = number_format.casefold()
    for marker, code in _CURRENCY_MARKERS:
        if marker in lowered:
            return code
    return None


def _looks_like_total_row(cells: dict[int, CellRecord]) -> bool:
    for record in cells.values():
        if _is_texty(record.value):
            words = normalize(str(record.value)).split()
            if words and (words[0] in _TOTAL_KEYWORDS or " ".join(words[:2]) in _TOTAL_KEYWORDS):
                return True
    return False


def _text_share(cells: dict[int, CellRecord]) -> float:
    if not cells:
        return 0.0
    texts = sum(1 for r in cells.values() if _is_texty(r.value) and r.formula is None)
    return texts / len(cells)


def _infer_column_type(
    records: list[CellRecord],
) -> tuple[ColumnType, str | None, str | None]:
    """Returns (type, dominant number format, currency code)."""
    values = [r.value for r in records if r.value is not None]
    if not values:
        return ColumnType.EMPTY, None, None

    n = len(values)
    dates = sum(1 for v in values if _is_date(v))
    bools = sum(
        1
        for v in values
        if isinstance(v, bool)
        or (isinstance(v, str) and normalize(v) in _BOOL_STRINGS)
    )
    numbers = sum(1 for v in values if _is_number(v))
    texts = sum(1 for v in values if _is_texty(v))

    formats = [r.number_format or "" for r in records if r.value is not None]
    dominant_format = max(set(formats), key=formats.count) if formats else None

    if dates / n >= 0.7:
        return ColumnType.DATE, dominant_format, None
    if bools / n >= 0.7:
        return ColumnType.BOOLEAN, dominant_format, None
    if numbers / n >= 0.7:
        currency = _currency_from_format(dominant_format)
        if dominant_format and "%" in dominant_format:
            return ColumnType.PERCENTAGE, dominant_format, None
        if currency:
            return ColumnType.CURRENCY, dominant_format, currency
        return ColumnType.NUMBER, dominant_format, None
    if texts / n >= 0.5:
        distinct = len({normalize(str(v)) for v in values if _is_texty(v)})
        if distinct <= 24 and distinct / n <= 0.6:
            return ColumnType.CATEGORICAL, dominant_format, None
        return ColumnType.TEXT, dominant_format, None
    return ColumnType.TEXT, dominant_format, None


def _likely_identifier(name: str, records: list[CellRecord], column_type: ColumnType) -> bool:
    values = [r.value for r in records if r.value is not None]
    if len(values) < 3:
        return False
    all_distinct = len({str(v) for v in values}) == len(values)
    name_says_id = "identifier" in concepts_of(name)
    if name_says_id and all_distinct:
        return True
    return bool(
        name_says_id and column_type in (ColumnType.NUMBER, ColumnType.TEXT)
    )


def _build_columns(
    rows: dict[int, dict[int, CellRecord]],
    col_indices: list[int],
    header_rows: list[int],
    body_rows: list[int],
) -> list[ColumnSchema]:
    """Column schemas for the given columns: names from the header rows,
    types/samples inferred from the body rows."""
    columns: list[ColumnSchema] = []
    for col_idx in col_indices:
        header_parts = [
            str(rows[hr][col_idx].value)
            for hr in header_rows
            if hr in rows and col_idx in rows[hr] and rows[hr][col_idx].value is not None
        ]
        name = (
            " / ".join(header_parts)
            if header_parts
            else f"Column {get_column_letter(col_idx)}"
        )
        body_records = [rows[r][col_idx] for r in body_rows if col_idx in rows[r]]
        column_type, number_format, currency = _infer_column_type(body_records)
        values = [r.value for r in body_records if r.value is not None]
        samples: list[str] = []
        for value in values:
            text = str(value)[:40]
            if text not in samples:
                samples.append(text)
            if len(samples) >= _SAMPLE_LIMIT:
                break
        columns.append(
            ColumnSchema(
                letter=get_column_letter(col_idx),
                index=col_idx,
                name=name,
                normalized_name=normalize(name),
                type=column_type,
                number_format=number_format or None,
                currency=currency,
                likely_identifier=_likely_identifier(name, body_records, column_type),
                missing_count=len(body_rows) - len(values),
                distinct_count=len({str(v) for v in values}),
                sample_values=samples,
            )
        )
    return columns


def _exact_table_schema(
    sheet: SheetInventory,
    rows: dict[int, dict[int, CellRecord]],
    block: list[int],
    block_cols: list[int],
) -> TableSchema | None:
    """Decision D14: an Excel Table object covering the block overrides the
    header/totals heuristics — its headerRowCount/totalsRowCount are exact
    metadata from the file. Returns None when no Table covers the block (the
    caller falls back to the heuristics)."""
    from openpyxl.utils.cell import range_boundaries

    for info in sheet.tables:
        try:
            c1, r1, c2, r2 = range_boundaries(info.ref)
        except ValueError:
            continue
        if c1 is None or r1 is None or c2 is None or r2 is None:
            continue
        if not (r1 <= block[0] and block[-1] <= r2):
            continue  # the Table does not cover the block's rows
        if c2 < block_cols[0] or c1 > block_cols[-1]:
            continue  # no column overlap
        header_count = max(info.header_row_count, 0)
        totals_count = max(info.totals_row_count, 0)
        header_rows = list(range(r1, r1 + header_count))
        data_start = r1 + header_count
        body_end = r2 - totals_count
        if body_end < data_start:
            continue  # header/totals only: nothing queryable
        total_rows = list(range(body_end + 1, r2 + 1))
        notes = [
            f"Excel Table '{info.name}': header/totals split is exact "
            "(from Excel Table metadata)."
        ]
        if total_rows:
            notes.append(
                "Row(s) "
                + ", ".join(str(r) for r in total_rows)
                + " are the Table's totals row(s) and are excluded from queries."
            )
        body_rows = [r for r in block if data_start <= r <= body_end]
        columns = _build_columns(rows, list(range(c1, c2 + 1)), header_rows, body_rows)
        return TableSchema(
            sheet_name=sheet.name,
            ref=str(info.ref),
            header_rows=header_rows,
            data_start_row=data_start,
            data_end_row=r2,
            row_count=body_end - data_start + 1,
            columns=columns,
            total_rows=total_rows,
            notes=notes,
        )
    return None


def _detect_block_tables(sheet: SheetInventory) -> list[TableSchema]:
    """Split the sheet into contiguous row blocks; each block that looks
    tabular (header row + >=2 data rows) becomes a TableSchema."""
    rows: dict[int, dict[int, CellRecord]] = {}
    for record in sheet.cells.values():
        rows.setdefault(record.row, {})[record.column] = record
    if not rows:
        return []

    # contiguous row blocks (a fully blank row separates tables)
    blocks: list[list[int]] = []
    for row_idx in sorted(rows):
        if blocks and row_idx == blocks[-1][-1] + 1:
            blocks[-1].append(row_idx)
        else:
            blocks.append([row_idx])

    tables: list[TableSchema] = []
    for block in blocks:
        block_cols = sorted({c for r in block for c in rows[r]})

        # D14: an Excel Table object covering the block gives the exact
        # header/totals split; the heuristics below apply only without one.
        exact = _exact_table_schema(sheet, rows, block, block_cols)
        if exact is not None:
            tables.append(exact)
            continue

        if len(block) < _MIN_TABLE_ROWS + 1:  # header + at least 2 data rows
            continue
        notes: list[str] = []

        # Header detection: first row in the block that is text-dominant and
        # spans at least half of the block's width.
        width = len(block_cols)
        header_rows: list[int] = []
        for pos, row_idx in enumerate(block[:3]):  # look at the first 3 rows only
            cells = rows[row_idx]
            if len(cells) >= max(2, width // 2) and _text_share(cells) >= _HEADER_TEXT_SHARE:
                header_rows = [row_idx]
                if pos > 0:
                    notes.append("Title row(s) above the header were ignored.")
                break
        if not header_rows:
            continue  # no recognizable header -> not a table we can query

        header_row = header_rows[0]
        # Multi-row header: the next row is also text-dominant while actual
        # data below is not.
        next_row = header_row + 1
        after_next = header_row + 2
        if (
            next_row in rows
            and after_next in rows
            and _text_share(rows[next_row]) >= _HEADER_TEXT_SHARE
            and _text_share(rows[after_next]) < _HEADER_TEXT_SHARE
        ):
            header_rows.append(next_row)
            notes.append("Two header rows detected and merged; verify column names.")

        data_start = header_rows[-1] + 1
        data_rows = [r for r in block if r >= data_start]
        if len(data_rows) < _MIN_TABLE_ROWS:
            continue

        total_rows = [r for r in data_rows if _looks_like_total_row(rows[r])]
        body_rows = [r for r in data_rows if r not in total_rows]
        if len(body_rows) < _MIN_TABLE_ROWS:
            continue
        if total_rows:
            notes.append(
                "Row(s) "
                + ", ".join(str(r) for r in total_rows)
                + " look like totals and are excluded from queries."
            )

        columns = _build_columns(rows, block_cols, header_rows, body_rows)

        first_col = get_column_letter(block_cols[0])
        last_col = get_column_letter(block_cols[-1])
        tables.append(
            TableSchema(
                sheet_name=sheet.name,
                ref=f"{first_col}{header_row}:{last_col}{block[-1]}",
                header_rows=header_rows,
                data_start_row=data_start,
                data_end_row=block[-1],
                row_count=len(body_rows),
                columns=columns,
                total_rows=total_rows,
                notes=notes,
            )
        )
    return tables


def detect_workbook_schema(inventory: WorkbookInventory) -> WorkbookSchema:
    tables: list[TableSchema] = []
    sheets: list[SheetSchemaInfo] = []
    warnings: list[str] = []

    for sheet in inventory.sheets:
        sheets.append(
            SheetSchemaInfo(
                name=sheet.name,
                visibility=sheet.visibility,
                has_formulas=sheet.formula_count > 0,
                hidden_rows=len(sheet.hidden_rows),
                hidden_columns=len(sheet.hidden_columns),
            )
        )
        if sheet.visibility.value != "visible":
            warnings.append(
                f"Sheet '{sheet.name}' is {sheet.visibility.value}; its tables are "
                "included but easy to miss in Excel."
            )
        if sheet.merged_ranges:
            for table in _detect_block_tables(sheet):
                overlapping = [
                    m for m in sheet.merged_ranges if _merge_overlaps(m, table)
                ]
                if overlapping:
                    table.notes.append(
                        f"Merged range(s) {', '.join(overlapping[:3])} overlap this "
                        "table; header interpretation may be approximate."
                    )
                tables.append(table)
        else:
            tables.extend(_detect_block_tables(sheet))

    return WorkbookSchema(
        workbook_id=inventory.workbook_id,
        filename=inventory.filename,
        sheets=sheets,
        tables=tables,
        warnings=warnings,
    )


def _merge_overlaps(merged_ref: str, table: TableSchema) -> bool:
    """Rough row-range overlap between a merged range and a table."""
    try:
        from openpyxl.utils.cell import range_boundaries

        _, min_row, _, max_row = range_boundaries(merged_ref)
    except ValueError:
        return False
    header_min = min(table.header_rows) if table.header_rows else table.data_start_row
    return not (max_row < header_min or min_row > table.data_end_row)
