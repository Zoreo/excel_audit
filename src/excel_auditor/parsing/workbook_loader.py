"""Safe workbook loading.

Validates the container (zip bombs, path traversal, non-Excel files) before
handing the file to openpyxl, then loads the workbook twice: once for
formulas/styles and once with data_only=True for cached values.

Macros are never executed; their presence is only detected.
"""

from __future__ import annotations

import logging
import zipfile
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
from openpyxl.workbook.workbook import Workbook

from ..config import Settings, get_settings
from ..errors import WorkbookLoadError, WorkbookValidationError

logger = logging.getLogger(__name__)

_ZIP_MAGIC = b"PK\x03\x04"


@dataclass(frozen=True)
class ZipFacts:
    """Facts derived from the raw zip container without parsing sheet XML."""

    has_vba: bool = False
    has_data_connections: bool = False
    external_link_parts: int = 0
    entry_count: int = 0
    total_uncompressed: int = 0


@dataclass
class LoadedWorkbook:
    path: Path
    formulas: Workbook  # data_only=False: formulas + styles
    values: Workbook  # data_only=True: cached values
    zip_facts: ZipFacts = field(default_factory=ZipFacts)
    file_size: int = 0


def validate_container(path: Path, settings: Settings | None = None) -> ZipFacts:
    """Validate the file is a safe .xlsx/.xlsm container. Raises WorkbookValidationError."""
    settings = settings or get_settings()

    if not path.is_file():
        raise WorkbookValidationError("File does not exist.")

    with open(path, "rb") as fh:
        magic = fh.read(4)
    if magic != _ZIP_MAGIC:
        raise WorkbookValidationError("Not a valid .xlsx workbook (not a zip archive).")

    try:
        with zipfile.ZipFile(path) as zf:
            infos = zf.infolist()
            names = zf.namelist()
    except zipfile.BadZipFile as exc:
        raise WorkbookValidationError("Corrupted or invalid zip archive.") from exc

    if len(infos) > settings.max_zip_entries:
        raise WorkbookValidationError(
            f"Archive contains too many entries ({len(infos)} > {settings.max_zip_entries})."
        )

    for name in names:
        if name.startswith(("/", "\\")) or ".." in name:
            raise WorkbookValidationError("Archive contains unsafe entry paths.")

    total_uncompressed = sum(i.file_size for i in infos)
    total_compressed = max(1, sum(i.compress_size for i in infos))
    if total_uncompressed > settings.max_decompressed_bytes:
        raise WorkbookValidationError(
            "Workbook decompresses to "
            f"{total_uncompressed // (1024 * 1024)} MB, above the "
            f"{settings.max_decompressed_mb} MB limit."
        )
    if (
        total_uncompressed > 10 * 1024 * 1024
        and total_uncompressed / total_compressed > settings.max_zip_ratio
    ):
        raise WorkbookValidationError("Suspicious compression ratio; file rejected.")

    name_set = set(names)
    if "xl/workbook.xml" not in name_set or "[Content_Types].xml" not in name_set:
        raise WorkbookValidationError("Not an Excel workbook (missing workbook parts).")

    return ZipFacts(
        has_vba="xl/vbaProject.bin" in name_set,
        has_data_connections="xl/connections.xml" in name_set,
        external_link_parts=sum(
            1 for n in names if n.startswith("xl/externalLinks/") and n.endswith(".xml")
        ),
        entry_count=len(infos),
        total_uncompressed=total_uncompressed,
    )


def load_workbook_safe(path: Path, settings: Settings | None = None) -> LoadedWorkbook:
    """Validate and load a workbook. Raises WorkbookValidationError / WorkbookLoadError."""
    path = Path(path)
    facts = validate_container(path, settings)
    try:
        formulas_wb = openpyxl.load_workbook(path, data_only=False, keep_links=True)
        values_wb = openpyxl.load_workbook(path, data_only=True, keep_links=True)
    except WorkbookValidationError:
        raise
    except Exception as exc:  # openpyxl raises a wide variety of exceptions
        logger.warning("Failed to parse workbook %s: %s", path.name, type(exc).__name__)
        raise WorkbookLoadError(f"Workbook could not be parsed ({type(exc).__name__}).") from exc

    return LoadedWorkbook(
        path=path,
        formulas=formulas_wb,
        values=values_wb,
        zip_facts=facts,
        file_size=path.stat().st_size,
    )
