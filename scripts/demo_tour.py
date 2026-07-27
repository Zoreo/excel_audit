#!/usr/bin/env python3
"""Narrated tour of excel-auditor's trust guarantees (report schema v3).

Builds small workbooks that reproduce classic silent-spreadsheet-error
situations, then shows how the engine handles each one: confirm instead of
guess, refuse instead of answering the wrong question, detect what other
tools miss, and produce byte-identical reports you can diff.

Run:  .venv/bin/python scripts/demo_tour.py
"""

from __future__ import annotations

import hashlib
import tempfile
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.workbook.defined_name import DefinedName

from excel_auditor.llm import UnsupportedQuestionError
from excel_auditor.llm.rule_parser import RuleBasedIntentParser
from excel_auditor.query_service import answer_query, inspect_schema
from excel_auditor.reporting.json_report import to_json
from excel_auditor.services import audit_workbook, compare_workbooks

STAMP = datetime(2026, 7, 24, 12, 0, 0, tzinfo=UTC)


def _banner(title: str) -> None:
    print()
    print("=" * 72)
    print(title)
    print("=" * 72)


def _duplicate_header_workbook(path: Path) -> Path:
    """A bank-export classic: two physically different 'Amount' columns."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"
    for col, name in enumerate(["Category", "Amount", "Amount"], start=1):
        ws.cell(row=1, column=col, value=name).font = Font(bold=True)
    rows = [("rent", 100, 1000), ("payroll", 200, 2000),
            ("hardware", 700, 7000), ("consulting", 900, 9000)]
    for i, (category, gross, net) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=category)
        ws.cell(row=i, column=2, value=gross)
        ws.cell(row=i, column=3, value=net)
    wb.save(path)
    return path


def _model_v1(path: Path) -> Path:
    """A small model: named input driving a Summary, plus a self-loop bug."""
    wb = Workbook()
    inputs = wb.active
    inputs.title = "Inputs"
    inputs["A1"] = 0.05                       # growth rate, the named input
    for row in range(1, 8):
        inputs[f"B{row}"] = row * 100
        inputs[f"C{row}"] = f"=B{row}*(1+GrowthRate)"
    summary = wb.create_sheet("Summary")
    summary["A1"] = "Total"
    summary["B1"] = "=SUM(Inputs!C1:C7)"
    bug = wb.create_sheet("Scratch")
    for row in range(1, 11):
        bug[f"D{row}"] = row
    bug["D11"] = "=SUM(D1:D11)"               # includes itself: silent classic
    wb.defined_names.add(DefinedName("GrowthRate", attr_text="Inputs!$A$1"))
    wb.save(path)
    return path


def _model_v2(path: Path) -> Path:
    """v2 of the same model: 'Inputs' renamed to 'Assumptions' + one edit."""
    wb = Workbook()
    inputs = wb.active
    inputs.title = "Assumptions"
    inputs["A1"] = 0.08                       # the edit hidden by the rename
    for row in range(1, 8):
        inputs[f"B{row}"] = row * 100
        inputs[f"C{row}"] = f"=B{row}*(1+GrowthRate)"
    summary = wb.create_sheet("Summary")
    summary["A1"] = "Total"
    summary["B1"] = "=SUM(Assumptions!C1:C7)"
    bug = wb.create_sheet("Scratch")
    for row in range(1, 11):
        bug[f"D{row}"] = row
    bug["D11"] = "=SUM(D1:D11)"
    wb.defined_names.add(DefinedName("GrowthRate", attr_text="Assumptions!$A$1"))
    wb.save(path)
    return path


def tour_confirm_instead_of_guess(workdir: Path) -> None:
    _banner("1. Ambiguity is confirmed, never guessed "
            "(duplicate 'Amount' headers)")
    path = _duplicate_header_workbook(workdir / "transactions.xlsx")
    parser = RuleBasedIntentParser()
    schema = inspect_schema(path).workbook_schema
    query = parser.parse("What is the total amount?", schema)

    first = answer_query(path, query)
    print(f'Q: "What is the total amount?"  ->  status: {first.result.status.value}')
    for i, cand in enumerate(first.result.candidates, start=1):
        print(f"   choice {i}: {cand.column_name} ({cand.column_type})")
    for choice in (1, 2):
        answer = answer_query(path, query, choices=[choice])
        prov = answer.result.provenance
        print(f"   with choice {choice}: {answer.result.formatted_value}"
              f"  [column: {prov.value_column}, status: {answer.result.status.value}]")
    print("   -> the two columns hold 1,900 vs 19,000; a silent guess would")
    print("      have been 10x wrong while claiming to be verified.")


def tour_refuse_instead_of_wrong(workdir: Path) -> None:
    _banner("2. Threshold questions filter correctly — or refuse")
    path = workdir / "transactions.xlsx"
    parser = RuleBasedIntentParser()
    schema = inspect_schema(path).workbook_schema

    query = parser.parse("show rows where amount over 500", schema)
    answer = answer_query(path, query, choices=[1])
    prov = answer.result.provenance
    print('Q: "show rows where amount over 500"')
    print(f"   -> {prov.rows_included} of {prov.rows_total} rows"
          f"  [filter: {', '.join(prov.filters)}]")

    print('Q: "show rows where amount is over budget"  (no number to filter on)')
    try:
        parser.parse("show rows where amount is over budget", schema)
    except UnsupportedQuestionError as exc:
        print(f"   -> REFUSED: {str(exc).split('.')[0]}.")
    print("   -> the old failure mode was returning ALL rows as 'verified'.")


def tour_audit_catches_the_classics(workdir: Path) -> None:
    _banner("3. Audit: self-including SUM and every finding on the record")
    v1 = _model_v1(workdir / "model_v1.xlsx")
    report = audit_workbook(v1, generated_at=STAMP)
    circular = [f for f in report.findings if f.rule_id == "EA-CIR-001"]
    print(f"Findings: {len(report.findings)}  |  risk level: {report.risk_level}")
    for f in circular:
        loc = f.location
        print(f"   EA-CIR-001 at {loc.sheet_name}!{loc.coordinate}: {f.title}")
    print(f"   rules that failed to run: {report.failed_rules or 'none'}")
    print("   -> Scratch!D11 sums a range that includes itself — invisible in")
    print("      Excel unless iterative calculation is off; flagged here.")
    print("   -> if an analysis rule ever crashes, the report says so instead")
    print("      of silently looking clean (failed_rules + coverage warning).")


def tour_named_input_impact(workdir: Path) -> None:
    _banner("4. Compare: named-range input changes carry downstream impact")
    v1, v2 = workdir / "model_v1.xlsx", workdir / "model_v2.xlsx"
    _model_v2(v2)
    report = compare_workbooks(v1, v2, generated_at=STAMP)

    renames = [c for c in report.structural_changes
               if c.change_type.value == "sheet_renamed"]
    for change in renames:
        print(f"Structural: {change.description}")
        print(f"   details: {change.details}")
    edits = [c for c in report.cell_changes if c.change_type.value == "value_changed"]
    for change in edits:
        impact = change.downstream_impact
        reach = (f"{impact.transitive_dependent_count} downstream cells, "
                 f"sheets: {', '.join(impact.affected_sheets)}"
                 if impact else "n/a")
        print(f"Cell edit: {change.sheet_name}!{change.coordinate} "
              f"{change.old_value} -> {change.new_value} "
              f"[severity: {change.severity.value}]")
        print(f"   downstream impact: {reach}")
    noise = [c for c in report.cell_changes if c.sheet_name == "Summary"]
    print(f"Bogus formula changes on Summary from the rename: {len(noise)}")
    print("   -> the rename is inferred from content (details.inferred=true),")
    print("      the hidden 0.05 -> 0.08 edit is surfaced WITH its blast")
    print("      radius via the GrowthRate defined name, and the rename alone")
    print("      produces zero formula-change noise.")


def tour_deterministic_reports(workdir: Path) -> None:
    _banner("5. Reports are evidence: byte-identical, content-addressed")
    v1 = workdir / "model_v1.xlsx"
    a = to_json(audit_workbook(v1, generated_at=STAMP))
    b = to_json(audit_workbook(v1, generated_at=STAMP))
    wb_id = audit_workbook(v1, generated_at=STAMP).workbook.workbook_id
    sha = hashlib.sha256(v1.read_bytes()).hexdigest()
    print(f"Two audit runs byte-identical: {a == b}")
    print(f"workbook_id == sha256(file):   {wb_id == sha}")
    print(f"report_schema_version:         "
          f"{audit_workbook(v1, generated_at=STAMP).report_schema_version}")
    print("   -> re-run months later on the same file and diff the JSON:")
    print("      zero noise. The id proves WHICH file the report describes.")


def _months_model(path: Path, months: int) -> Path:
    """A monthly table: label, amount, running formula — plus a grand total."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Budget"
    ws["A1"], ws["B1"], ws["C1"] = "Month", "Amount", "Running"
    for i in range(1, months + 1):
        row = i + 1
        ws[f"A{row}"] = f"2026-{i:02d}"
        ws[f"B{row}"] = 100 * i
        ws[f"C{row}"] = f"=SUM(B$2:B{row})"
    ws[f"A{months + 2}"] = "Total"
    ws[f"B{months + 2}"] = f"=SUM(B2:B{months + 1})"
    wb.save(path)
    return path


def tour_row_insertion_collapse(workdir: Path) -> None:
    _banner("6. A row inserted mid-table is ONE change, not a wall of noise")
    v10 = _months_model(workdir / "budget_10m.xlsx", 10)
    # 11-month version = same model with one month inserted mid-table.
    wb = Workbook()
    ws = wb.active
    ws.title = "Budget"
    ws["A1"], ws["B1"], ws["C1"] = "Month", "Amount", "Running"
    months = [f"2026-{i:02d}" for i in range(1, 11)]
    months.insert(5, "2026-05b")              # the inserted row
    for idx, label in enumerate(months, start=2):
        base = (idx - 1) if idx <= 6 else (idx - 2 if label != "2026-05b" else 55)
        ws[f"A{idx}"] = label
        ws[f"B{idx}"] = 100 * base if label != "2026-05b" else 550
        ws[f"C{idx}"] = f"=SUM(B$2:B{idx})"
    ws["A13"] = "Total"
    ws["B13"] = "=SUM(B2:B12)"                # total range grew: real change
    v11 = workdir / "budget_11m.xlsx"
    wb.save(v11)

    report = compare_workbooks(v10, v11, generated_at=STAMP)
    inserted = [c for c in report.structural_changes
                if c.change_type.value == "rows_inserted"]
    cell_adds = [c for c in report.cell_changes
                 if c.change_type.value in ("value_added", "formula_added")]
    formula_edits = [c for c in report.cell_changes
                     if c.change_type.value == "formula_changed"]
    for change in inserted:
        print(f"Structural: {change.description}")
        print(f"   details: start_row={change.details.get('start_row')}, "
              f"count={change.details.get('count')}")
    print(f"Cell-level added-value noise from the shifted rows: {len(cell_adds)}")
    for change in formula_edits:
        print(f"Real change kept: {change.sheet_name}!{change.coordinate} "
              f"{change.old_formula} -> {change.new_formula}")
    print("   -> before schema v3 this diff was dozens of value/formula 'adds';")
    print("      now it is one rows_inserted plus the genuinely grown total.")


def _invoice_with_drift(path: Path) -> Path:
    """Line items with sub-cent residue and a hand-corrected total."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Фактури"
    ws["A1"], ws["B1"] = "Услуга", "Сума"
    items = [("Консултации", 1234.564), ("Поддръжка", 2345.333), ("Лицензи", 876.108)]
    for row, (label, amount) in enumerate(items, start=2):
        ws[f"A{row}"] = label
        ws[f"B{row}"] = amount
        ws[f"B{row}"].number_format = '#,##0.00 "лв"'
    ws["A5"] = "Общо"
    ws["B5"] = 4456.01                        # typed in to "fix" the total
    ws["B5"].number_format = '#,##0.00 "лв"'
    wb.save(path)
    return path


def tour_rounding_drift(workdir: Path) -> None:
    _banner("7. The invoice that's off by a cent — found, explained, cells named")
    path = _invoice_with_drift(workdir / "invoice.xlsx")
    report = audit_workbook(path, generated_at=STAMP)
    for f in (f for f in report.findings if f.rule_id == "EA-RND-001"):
        loc = f.location
        ev = f.evidence
        print(f"EA-RND-001 at {loc.sheet_name}!{loc.coordinate}: {f.title}")
        print(f"   displayed components {ev['displayed_components_sum']} vs "
              f"displayed total {ev['displayed_total']} "
              f"(drift {ev['drift']} {ev.get('currency', '')})")
        for res in ev["residue_cells"]:
            print(f"   residue: {res['cell']} stores {res['stored']} "
                  f"but displays {res['displayed']}")
    print("   -> the printed column really is off by one stotinka; the guilty")
    print("      cells carry sub-display precision that SUM sees and the eye")
    print("      doesn't. The fix is a ROUND() policy, not a typed-over total.")


def main() -> None:
    print("excel-auditor — trust-guarantee tour (report schema v3)")
    with tempfile.TemporaryDirectory(prefix="excel-auditor-tour-") as tmp:
        workdir = Path(tmp)
        tour_confirm_instead_of_guess(workdir)
        tour_refuse_instead_of_wrong(workdir)
        tour_audit_catches_the_classics(workdir)
        tour_named_input_impact(workdir)
        tour_deterministic_reports(workdir)
        tour_row_insertion_collapse(workdir)
        tour_rounding_drift(workdir)
    print()
    print("Done. Full reports: excel-auditor audit/compare/ask <file> "
          "(HTML + JSON with provenance).")


if __name__ == "__main__":
    main()
