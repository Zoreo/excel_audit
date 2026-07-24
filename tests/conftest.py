"""Shared fixtures: generated demo workbooks and a tiny workbook builder."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook
from openpyxl.styles import Font

from excel_auditor.analysis.workbook_inventory import inventory_from_path
from excel_auditor.demo import generate_demo_workbooks
from excel_auditor.models import WorkbookInventory


@pytest.fixture(scope="session")
def demo_paths(tmp_path_factory: pytest.TempPathFactory) -> tuple[Path, Path]:
    directory = tmp_path_factory.mktemp("demo")
    return generate_demo_workbooks(directory)


@pytest.fixture(scope="session")
def old_inventory(demo_paths: tuple[Path, Path]) -> WorkbookInventory:
    return inventory_from_path(demo_paths[0], workbook_id="old")


@pytest.fixture(scope="session")
def new_inventory(demo_paths: tuple[Path, Path]) -> WorkbookInventory:
    return inventory_from_path(demo_paths[1], workbook_id="new")


def make_workbook(
    path: Path,
    sheets: dict[str, dict[str, Any]],
    *,
    hidden_sheets: tuple[str, ...] = (),
) -> Path:
    """Build a small workbook from {sheet: {coordinate: value}}. Strings that
    start with '=' become formulas."""
    wb = Workbook()
    default = wb.active
    for index, (name, cells) in enumerate(sheets.items()):
        ws = default if index == 0 else wb.create_sheet()
        ws.title = name
        for coordinate, value in cells.items():
            ws[coordinate] = value
        if name in hidden_sheets:
            ws.sheet_state = "hidden"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


# ---------------------------------------------------------------- sales data

# Deliberately crafted numbers: 2025 revenue sums to 628,400 (the product
# spec's example) with 2 blank values; 2024 sums to 400,000.
SALES_ROWS = [
    # (date, customer, region, revenue, net_revenue, paid)
    (datetime(2024, 3, 15), "Alfa", "София", 150000, 135000, True),
    (datetime(2024, 6, 20), "Beta", "Пловдив", 130000, 117000, True),
    (datetime(2024, 11, 5), "Alfa", "София", 120000, 108000, False),
    (datetime(2025, 1, 15), "Alfa", "София", 100000, 90000, True),
    (datetime(2025, 2, 10), "Beta", "Пловдив", 120000, 108000, True),
    (datetime(2025, 3, 12), "Gama", "Варна", 150000, 135000, False),
    (datetime(2025, 4, 18), "Alfa", "София", 130000, 117000, True),
    (datetime(2025, 5, 22), "Delta", "София", 128400, 115560, True),
    (datetime(2025, 6, 15), "Epsilon", "Варна", None, None, False),
    (datetime(2025, 7, 1), "Beta", "Пловдив", None, None, False),
]

_HEADER_SETS = {
    "bg": ["Дата", "Клиент", "Регион", "Оборот", "Нетен оборот", "Платено"],
    "en": ["Date", "Customer", "Region", "Revenue", "Net Revenue", "Paid"],
    "translit": ["Data", "Klient", "Region", "Oborot", "Neten oborot", "Plateno"],
}


def make_sales_workbook(
    path: Path,
    *,
    headers: str = "bg",
    net_column: bool = True,
    subtotal: bool = True,
    second_date: bool = False,
    forecast_columns: bool = False,
) -> Path:
    """Realistic sales table with currency formats, blanks and a subtotal row."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    names = list(_HEADER_SETS[headers])
    if not net_column:
        names = names[:4] + names[5:]
    if second_date:
        names.append("Дата на плащане" if headers == "bg" else "Payment Date")
    if forecast_columns:
        names += ["Revenue Actual", "Revenue Forecast"]
    for col, name in enumerate(names, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = Font(bold=True)

    row_idx = 2
    for entry in SALES_ROWS:
        stamp, customer, region, revenue, net, paid = entry
        values = [stamp, customer, region, revenue]
        if net_column:
            values.append(net)
        values.append(paid)
        if second_date:
            values.append(datetime(stamp.year, stamp.month, 28))
        if forecast_columns:
            values += [revenue, (revenue or 0) * 1.1 if revenue else None]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            name = names[col - 1].casefold()
            if isinstance(value, datetime):
                cell.number_format = "yyyy-mm-dd"
            elif isinstance(value, (int, float)) and not isinstance(value, bool):
                cell.number_format = "€#,##0.00"
        row_idx += 1

    if subtotal:
        label = "Общо" if headers == "bg" else "Total"
        ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        revenue_col = 4
        total = sum(r[3] for r in SALES_ROWS if r[3] is not None)
        cell = ws.cell(row=row_idx, column=revenue_col, value=total)
        cell.number_format = "€#,##0.00"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def make_contracts_workbook(path: Path) -> Path:
    """Contracts with deadlines around a 2026-07-24 reference date."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Договори"
    for col, name in enumerate(["Договор", "Краен срок", "Стойност"], start=1):
        ws.cell(row=1, column=col, value=name).font = Font(bold=True)
    rows = [
        ("Договор А", datetime(2026, 7, 30), 12000),
        ("Договор Б", datetime(2026, 8, 10), 8000),
        ("Договор В", datetime(2026, 9, 15), 20000),
        ("Договор Г", datetime(2025, 12, 1), 5000),
    ]
    for i, (name, due, amount) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=name)
        due_cell = ws.cell(row=i, column=2, value=due)
        due_cell.number_format = "yyyy-mm-dd"
        ws.cell(row=i, column=3, value=amount)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


@pytest.fixture(scope="session")
def sales_bg(tmp_path_factory: pytest.TempPathFactory) -> Path:
    return make_sales_workbook(
        tmp_path_factory.mktemp("sales") / "sales_bg.xlsx", headers="bg"
    )


@pytest.fixture(scope="session")
def sales_bg_simple(tmp_path_factory: pytest.TempPathFactory) -> Path:
    """Single revenue column -> no gross/net ambiguity."""
    return make_sales_workbook(
        tmp_path_factory.mktemp("sales_simple") / "sales_simple.xlsx",
        headers="bg",
        net_column=False,
    )


@pytest.fixture(scope="session")
def sales_en(tmp_path_factory: pytest.TempPathFactory) -> Path:
    return make_sales_workbook(
        tmp_path_factory.mktemp("sales_en") / "sales_en.xlsx",
        headers="en",
        net_column=False,
    )


@pytest.fixture(scope="session")
def contracts(tmp_path_factory: pytest.TempPathFactory) -> Path:
    return make_contracts_workbook(
        tmp_path_factory.mktemp("contracts") / "contracts.xlsx"
    )
