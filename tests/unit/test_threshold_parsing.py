"""QA-002 regressions: numeric thresholds must be extracted or the question
refused — never silently answered unfiltered."""

import pytest

from excel_auditor.llm import UnsupportedQuestionError
from excel_auditor.llm.rule_parser import RuleBasedIntentParser
from excel_auditor.models.query import (
    FilterOperator,
    QueryOperation,
    ResultStatus,
)
from excel_auditor.query_service import answer_query, inspect_schema


@pytest.fixture(scope="module")
def parser() -> RuleBasedIntentParser:
    return RuleBasedIntentParser()


@pytest.fixture()
def amounts_workbook(tmp_path):
    """Customer | Amount with amounts 100/200/700/900 (two rows above 500)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    path = tmp_path / "amounts.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    for col, name in enumerate(["Customer", "Amount"], start=1):
        ws.cell(row=1, column=col, value=name).font = Font(bold=True)
    rows = [("Alfa", 100), ("Beta", 200), ("Gama", 700), ("Delta", 900)]
    for i, (customer, amount) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=customer)
        ws.cell(row=i, column=2, value=amount)
    wb.save(path)
    return path


def _schema(path):
    return inspect_schema(path).workbook_schema


# ------------------------------------------------------------ parse level


def test_list_rows_threshold_builds_filter(parser, amounts_workbook):
    query = parser.parse("show rows where amount over 500", _schema(amounts_workbook))
    assert query.operation == QueryOperation.LIST_ROWS
    assert len(query.filters) == 1
    flt = query.filters[0]
    assert flt.column == "amount"
    assert flt.operator == FilterOperator.GREATER_THAN
    assert flt.value == 500


@pytest.mark.parametrize(
    "question",
    [
        "show rows where amount exceeds 500",
        "show rows where amount above 500",
        "show rows where amount more than 500",
        "покажи редове където сума над 500",
        "покажи редове където сума повече от 500",
    ],
)
def test_threshold_keyword_variants(parser, amounts_workbook, question):
    query = parser.parse(question, _schema(amounts_workbook))
    assert query.operation == QueryOperation.LIST_ROWS
    assert [
        (f.column, f.operator, f.value) for f in query.filters
    ] == [("amount", FilterOperator.GREATER_THAN, 500)]


def test_threshold_with_thousands_separator(parser, amounts_workbook):
    query = parser.parse(
        "show rows where amount over 10 000", _schema(amounts_workbook)
    )
    assert query.filters[0].value == 10000


def test_threshold_keyword_without_number_refuses(parser, amounts_workbook):
    with pytest.raises(UnsupportedQuestionError):
        parser.parse("show rows where amount over budget", _schema(amounts_workbook))


def test_threshold_without_metric_refuses(parser, amounts_workbook):
    with pytest.raises(UnsupportedQuestionError):
        parser.parse("show rows where flurbs over 500", _schema(amounts_workbook))


def test_aggregate_threshold_builds_filter(parser, amounts_workbook):
    query = parser.parse(
        "how many customers have amount over 500", _schema(amounts_workbook)
    )
    assert query.operation == QueryOperation.AGGREGATE
    assert [
        (f.column, f.operator, f.value) for f in query.filters
    ] == [("amount", FilterOperator.GREATER_THAN, 500)]


def test_aggregate_threshold_without_number_refuses(parser, amounts_workbook):
    with pytest.raises(UnsupportedQuestionError):
        parser.parse(
            "how many customers have amount over the plan", _schema(amounts_workbook)
        )


def test_year_filter_not_hijacked_as_threshold(parser, sales_bg):
    query = parser.parse("What was total revenue in 2024?", _schema(sales_bg))
    assert [
        (f.column, f.operator, f.value) for f in query.filters
    ] == [("__date__", FilterOperator.YEAR_EQUALS, 2024)]


def test_year_over_year_is_not_a_threshold(parser, sales_bg):
    # comparison phrasing must not trip the threshold refusal
    query = parser.parse("show revenue year over year", _schema(sales_bg))
    assert query.operation == QueryOperation.AGGREGATE
    assert query.filters == []


def test_deadline_days_not_swallowed_as_threshold(parser, sales_bg):
    query = parser.parse("Колко договора изтичат до 30 дни?", _schema(sales_bg))
    assert query.operation == QueryOperation.DUE_WITHIN
    assert query.horizon_days == 30
    assert query.filters == []


# ------------------------------------------------------------- end to end


def test_list_rows_threshold_end_to_end(amounts_workbook):
    parser = RuleBasedIntentParser()
    query = parser.parse(
        "show rows where amount over 500", _schema(amounts_workbook)
    )
    report = answer_query(amounts_workbook, query, exact_columns=False)
    result = report.result
    assert result.status == ResultStatus.VERIFIED
    assert result.value == 2
    assert {row["Customer"] for row in result.rows} == {"Gama", "Delta"}
    assert result.provenance.filters == ["Amount greater than 500.0"]


def test_aggregate_threshold_end_to_end_counts_filtered(amounts_workbook):
    parser = RuleBasedIntentParser()
    query = parser.parse(
        "how many customers have amount over 500", _schema(amounts_workbook)
    )
    report = answer_query(amounts_workbook, query, exact_columns=False)
    result = report.result
    assert result.status == ResultStatus.VERIFIED
    assert result.value == 2  # never the unfiltered count of 4
    assert result.provenance.filters == ["Amount greater than 500.0"]
