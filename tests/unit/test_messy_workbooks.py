"""Realistic, messy workbook scenarios.

The generated demo model is deliberately clean; these tests confirm the engine
behaves sensibly on the layouts real accounting files actually have: spacer
rows, several blocks per sheet, merged headers, horizontal copies, Cyrillic
sheet names, inserted/reordered rows, duplicated blocks, intentional override
cells, named ranges, and modern functions incl. structured table references.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table

from conftest import make_workbook
from excel_auditor.analysis.dependency_graph import DependencyGraph
from excel_auditor.analysis.pattern_detection import detect_pattern_anomalies
from excel_auditor.analysis.workbook_inventory import inventory_from_path
from excel_auditor.models.enums import SEVERITY_ORDER, Severity
from excel_auditor.services import audit_workbook, compare_workbooks


def _pattern_findings(report):
    return [f for f in report.findings if f.rule_id.startswith("EA-PAT")]


# ---------------------------------------------------------------- layout noise


def test_blank_spacer_rows_are_not_missing_formulas(tmp_path: Path):
    """A fully blank row between two halves of the same block is layout."""
    cells = {}
    for r in list(range(2, 7)) + list(range(8, 13)):  # row 7 entirely blank
        cells[f"A{r}"] = r
        cells[f"B{r}"] = f"=A{r}*2"
    path = make_workbook(tmp_path / "spacer.xlsx", {"Data": cells})
    anomalies = detect_pattern_anomalies(inventory_from_path(path))
    assert anomalies == []


def test_partially_blank_row_is_still_flagged(tmp_path: Path):
    """If the row has other data, a hole in the formula column is suspicious."""
    cells = {}
    for r in range(2, 9):
        cells[f"A{r}"] = r
        if r != 5:
            cells[f"B{r}"] = f"=A{r}*2"
    path = make_workbook(tmp_path / "hole.xlsx", {"Data": cells})
    anomalies = detect_pattern_anomalies(inventory_from_path(path))
    assert [(a.coordinate, a.kind) for a in anomalies] == [("B5", "missing_formula")]


def test_multiple_blocks_on_one_sheet(tmp_path: Path):
    """Two unrelated blocks with different patterns must not cross-contaminate."""
    cells = {"A1": "Salaries"}
    for r in range(2, 7):
        cells[f"A{r}"] = r * 100
        cells[f"B{r}"] = f"=A{r}*1.2"
    cells["A9"] = "Bonuses"
    for r in range(10, 15):
        cells[f"A{r}"] = r * 10
        cells[f"B{r}"] = f"=A{r}+50"
    path = make_workbook(tmp_path / "blocks.xlsx", {"Costs": cells})
    anomalies = detect_pattern_anomalies(inventory_from_path(path))
    assert anomalies == []


def test_duplicated_blocks_stay_clean(tmp_path: Path):
    """The same block copied lower on the sheet (quarterly layout)."""
    cells = {}
    for start in (2, 10, 18):
        for offset in range(5):
            r = start + offset
            cells[f"A{r}"] = offset + 1
            cells[f"B{r}"] = f"=A{r}*2"
    path = make_workbook(tmp_path / "duplicated.xlsx", {"Quarters": cells})
    anomalies = detect_pattern_anomalies(inventory_from_path(path))
    assert anomalies == []


def test_merged_headers(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет"
    ws["A1"] = "Годишен отчет 2025"
    ws.merge_cells("A1:D1")
    for col, header in zip("ABCD", ["Месец", "Бройки", "Цена", "Приход"], strict=True):
        ws[f"{col}2"] = header
    for r in range(3, 9):
        ws[f"A{r}"] = r - 2
        ws[f"B{r}"] = 100 + r
        ws[f"C{r}"] = 9.99
        ws[f"D{r}"] = f"=B{r}*C{r}"
    path = tmp_path / "merged.xlsx"
    wb.save(path)

    report = audit_workbook(path)
    inventory = inventory_from_path(path)
    assert "A1:D1" in inventory.sheets[0].merged_ranges
    assert _pattern_findings(report) == []


def test_formulas_copied_horizontally_with_override(tmp_path: Path):
    """Horizontal pattern with one hardcoded override in the middle."""
    cells = {}
    for i, col in enumerate("BCDEFG"):
        cells[f"{col}1"] = 100 + i
        cells[f"{col}2"] = f"={col}1*1.1" if col != "E" else 999
    path = make_workbook(tmp_path / "horizontal.xlsx", {"Data": cells})
    anomalies = detect_pattern_anomalies(inventory_from_path(path))
    hits = [a for a in anomalies if a.coordinate == "E2"]
    assert len(hits) == 1
    assert hits[0].kind == "overwritten_constant"
    assert hits[0].orientation == "row"


# ------------------------------------------------------------ Bulgarian names


def test_bulgarian_sheet_names(tmp_path: Path):
    path = make_workbook(
        tmp_path / "bg.xlsx",
        {
            "Приходи": {"A1": "Продажби", "B1": 1000, "B2": 1200, "B3": "=SUM(B1:B2)"},
            "Разходи": {"A1": "Заплати", "B1": 700, "B2": 100, "B3": "=SUM(B1:B2)"},
            "Обобщение": {"A1": "Печалба", "B1": "='Приходи'!B3-'Разходи'!B3"},
        },
    )
    inventory = inventory_from_path(path)
    graph = DependencyGraph.build(inventory)
    assert ("Обобщение", "B1") in graph.dependents[("Приходи", "B3")]
    assert ("Обобщение", "B1") in graph.dependents[("Разходи", "B3")]

    report = audit_workbook(path)
    assert _pattern_findings(report) == []
    assert not [f for f in report.findings if f.rule_id == "EA-EXT-001"]


def test_unquoted_cyrillic_sheet_reference(tmp_path: Path):
    path = make_workbook(
        tmp_path / "bg2.xlsx",
        {
            "Данни": {"B2": 42},
            "Изход": {"A1": "=Данни!B2*2"},
        },
    )
    graph = DependencyGraph.build(inventory_from_path(path))
    assert ("Изход", "A1") in graph.dependents[("Данни", "B2")]


# ------------------------------------------------------- structural evolution


def _model_with_rows(path: Path, months: int, *, total_label_row_offset: int = 1) -> Path:
    cells = {"A1": "Month", "B1": "Units", "C1": "Price", "D1": "Revenue"}
    for i in range(months):
        r = i + 2
        cells[f"A{r}"] = i + 1
        cells[f"B{r}"] = 100 + i * 3
        cells[f"C{r}"] = 25
        cells[f"D{r}"] = f"=B{r}*C{r}"
    total_row = months + 1 + total_label_row_offset
    cells[f"A{total_row}"] = "Total"
    cells[f"D{total_row}"] = f"=SUM(D2:D{months + 1})"
    return make_workbook(path, {"Data": cells})


def test_inserted_row(tmp_path: Path):
    """Row inserted mid-table: engine completes, and the grown v2 stays
    pattern-clean (no false positives from the shift)."""
    v1 = _model_with_rows(tmp_path / "v1.xlsx", months=10)
    v2 = _model_with_rows(tmp_path / "v2.xlsx", months=11)

    report = compare_workbooks(v1, v2)
    assert report.summary.total_cell_changes > 0

    audit = audit_workbook(v2)
    assert _pattern_findings(audit) == []
    assert not [f for f in audit.findings if f.rule_id == "EA-RNG-001"]


def test_reordered_rows(tmp_path: Path):
    """Two data rows swapped: value changes only, nothing high-severity."""
    def build(path, units):
        cells = {}
        for i, u in enumerate(units):
            r = i + 2
            cells[f"B{r}"] = u
            cells[f"D{r}"] = f"=B{r}*2"
        cells["A8"] = "Total"
        cells["D8"] = "=SUM(D2:D7)"
        return make_workbook(path, {"Data": cells})

    v1 = build(tmp_path / "v1.xlsx", [10, 20, 30, 40, 50, 60])
    v2 = build(tmp_path / "v2.xlsx", [20, 10, 30, 40, 50, 60])

    report = compare_workbooks(v1, v2)
    assert {c.change_type.value for c in report.cell_changes} == {"value_changed"}
    assert all(
        SEVERITY_ORDER[i.severity] <= SEVERITY_ORDER[Severity.MEDIUM]
        for i in report.review_items
    )
    audit = audit_workbook(v2)
    assert _pattern_findings(audit) == []


def test_intentional_override_is_flagged_for_review(tmp_path: Path):
    """An intentionally hardcoded override cell IS flagged - as a review
    prompt, high confidence, suggesting documentation rather than 'error'."""
    cells = {}
    for r in range(2, 9):
        cells[f"A{r}"] = r
        cells[f"B{r}"] = f"=A{r}*100" if r != 5 else 4200  # agreed override
    path = make_workbook(tmp_path / "override.xlsx", {"Plan": cells})
    report = audit_workbook(path)
    overrides = [f for f in report.findings if f.rule_id == "EA-PAT-001"]
    assert len(overrides) == 1
    assert overrides[0].location is not None
    assert overrides[0].location.coordinate == "B5"
    assert "document" in overrides[0].suggested_action.lower()


# ----------------------------------------------------- names & modern formulas


def test_formulas_using_named_ranges(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Модел"
    ws["E1"] = 0.05
    for r in range(2, 8):
        ws[f"A{r}"] = r * 10
        ws[f"B{r}"] = f"=A{r}*(1+Growth)"
    wb.defined_names["Growth"] = DefinedName("Growth", attr_text="Модел!$E$1")
    path = tmp_path / "named.xlsx"
    wb.save(path)

    inventory = inventory_from_path(path)
    sheet = inventory.sheets[0]
    # named ranges normalize case-insensitively, so copies compare equal
    assert sheet.cells["B2"].normalized_formula == sheet.cells["B7"].normalized_formula

    report = audit_workbook(path)
    assert _pattern_findings(report) == []
    assert not [f for f in report.findings if f.rule_id == "EA-RNG-002"]
    assert not [f for f in report.findings if f.rule_id == "EA-EXT-001"]


def test_modern_functions_and_structured_references(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    for col, header in zip("ABC", ["Product", "Qty", "Amount"], strict=True):
        ws[f"{col}1"] = header
    data = [("Widget", 10, 250), ("Gadget", 5, 400), ("Widget", 3, 75), ("Doohickey", 8, 320)]
    for i, (product, qty, amount) in enumerate(data):
        r = i + 2
        ws[f"A{r}"] = product
        ws[f"B{r}"] = qty
        ws[f"C{r}"] = amount
    ws.add_table(Table(displayName="SalesTbl", ref="A1:C5"))

    ws["E1"] = "Widget"
    ws["E2"] = "=SUMIFS(C2:C5,A2:A5,E1)"
    ws["E3"] = '=INDEX(C2:C5,MATCH("Gadget",A2:A5,0))'
    ws["E4"] = "=XLOOKUP(E1,A2:A5,C2:C5)"
    ws["E5"] = '=IF(E2>100,"over","under")'
    ws["E6"] = "=SUM(SalesTbl[Amount])"
    ws["E7"] = "=SUM(SalesTbl[[#All],[Amount]])"
    path = tmp_path / "modern.xlsx"
    wb.save(path)

    report = audit_workbook(path)  # must not raise
    fired = {f.rule_id for f in report.findings}
    # structured refs and table names must NOT read as external dependencies
    assert "EA-EXT-001" not in fired
    assert "EA-REF-001" not in fired
    assert "EA-VOL-001" not in fired
    assert _pattern_findings(report) == []
    # the Table object itself is inventoried
    inventory = inventory_from_path(path)
    assert inventory.sheets[0].tables and inventory.sheets[0].tables[0].name == "SalesTbl"


def test_sumifs_copied_down_stays_consistent(tmp_path: Path):
    cells = {"F1": "Widget", "F2": "Gadget", "F3": "Doohickey", "F4": "Sprocket"}
    for r in range(1, 5):
        cells[f"G{r}"] = f"=SUMIFS($C$2:$C$50,$A$2:$A$50,F{r})"
    for r in range(2, 12):
        cells[f"A{r}"] = "Widget"
        cells[f"C{r}"] = r * 10
    path = make_workbook(tmp_path / "sumifs.xlsx", {"Pivot": cells})
    inventory = inventory_from_path(path)
    sheet = inventory.sheets[0]
    assert sheet.cells["G1"].normalized_formula == sheet.cells["G4"].normalized_formula
    assert detect_pattern_anomalies(inventory) == []
