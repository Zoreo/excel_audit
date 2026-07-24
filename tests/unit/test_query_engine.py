from datetime import date

import pytest

from excel_auditor.analysis.query import ResolvedQuery, execute_query, load_table_frame
from excel_auditor.analysis.schema import detect_workbook_schema
from excel_auditor.analysis.workbook_inventory import inventory_from_path
from excel_auditor.models.query import (
    AggregateFunction,
    FilterOperator,
    PeriodComparison,
    QueryFilter,
    QueryOperation,
    ResultStatus,
    SpreadsheetQuery,
)
from excel_auditor.query_service import answer_query


def _query(**kwargs) -> SpreadsheetQuery:
    return SpreadsheetQuery(**kwargs)


def _year_filter(year: int) -> QueryFilter:
    return QueryFilter(column="__date__", operator=FilterOperator.YEAR_EQUALS, value=year)


def test_sum_with_year_filter_and_provenance(sales_bg):
    report = answer_query(
        sales_bg,
        _query(
            operation=QueryOperation.AGGREGATE,
            function=AggregateFunction.SUM,
            requested_metric="Оборот",
            filters=[_year_filter(2025)],
        ),
        exact_columns=True,
    )
    result = report.result
    assert result.status == ResultStatus.REVIEW_RECOMMENDED  # blanks excluded
    assert result.value == 628400
    assert result.formatted_value == "€628,400"
    p = result.provenance
    assert p.sheet == "Sales"
    assert p.value_column == "Оборот"
    assert p.date_column == "Дата"
    assert p.filters == ["year(Дата) = 2025"]
    assert p.rows_included == 7
    assert p.rows_excluded_blank == 2
    assert p.rows_excluded_total_rows == 1
    assert p.currency == "EUR"
    assert any("total/subtotal" in a for a in p.assumptions)


def test_subtotal_row_not_double_counted(sales_bg):
    report = answer_query(
        sales_bg,
        _query(function=AggregateFunction.SUM, requested_metric="Оборот"),
        exact_columns=True,
    )
    # all data rows, subtotal excluded: 400000 + 628400
    assert report.result.value == 1028400


def test_average_min_max_median(sales_bg_simple):
    def run(function):
        return answer_query(
            sales_bg_simple,
            _query(
                function=function,
                requested_metric="Оборот",
                filters=[_year_filter(2025)],
            ),
            exact_columns=True,
        ).result.value

    assert run(AggregateFunction.AVERAGE) == 125680
    assert run(AggregateFunction.MINIMUM) == 100000
    assert run(AggregateFunction.MAXIMUM) == 150000
    assert run(AggregateFunction.MEDIAN) == 128400


def test_distinct_count(sales_bg):
    report = answer_query(
        sales_bg,
        _query(function=AggregateFunction.DISTINCT_COUNT, requested_metric="Клиент"),
        exact_columns=True,
    )
    assert report.result.value == 5


def test_group_by_region(sales_bg):
    report = answer_query(
        sales_bg,
        _query(
            function=AggregateFunction.SUM,
            requested_metric="Оборот",
            group_by=["Регион"],
        ),
        exact_columns=True,
    )
    groups = {row.key["Регион"]: row.value for row in report.result.groups}
    assert groups["София"] == 628400
    assert groups["Пловдив"] == 250000
    assert groups["Варна"] == 150000


def test_group_by_month(sales_bg_simple):
    report = answer_query(
        sales_bg_simple,
        _query(
            function=AggregateFunction.SUM,
            requested_metric="Оборот",
            group_by=["month"],
            filters=[_year_filter(2025)],
        ),
        exact_columns=True,
    )
    groups = {row.key["Дата"]: row.value for row in report.result.groups}
    assert groups["2025-01"] == 100000
    assert groups["2025-05"] == 128400
    assert any("Grouped by month" in a for a in report.result.provenance.assumptions)


def test_period_comparison(sales_bg_simple):
    report = answer_query(
        sales_bg_simple,
        _query(
            operation=QueryOperation.COMPARE_PERIODS,
            function=AggregateFunction.SUM,
            requested_metric="Оборот",
            period_comparison=PeriodComparison(period_a=2024, period_b=2025),
        ),
        exact_columns=True,
    )
    comparison = report.result.comparison
    assert comparison["value_a"] == 400000
    assert comparison["value_b"] == 628400
    assert comparison["change"] == 228400
    assert comparison["pct_change"] == 57.1


def test_filter_greater_than_lists_rows(sales_bg_simple):
    report = answer_query(
        sales_bg_simple,
        _query(
            operation=QueryOperation.LIST_ROWS,
            filters=[
                QueryFilter(
                    column="Оборот",
                    operator=FilterOperator.GREATER_THAN,
                    value=125000,
                )
            ],
        ),
        exact_columns=True,
    )
    # 150000+130000 (2024) and 150000+130000+128400 (2025)
    assert report.result.value == 5
    assert len(report.result.rows) == 5


def test_next_deadline(contracts):
    report = answer_query(
        contracts,
        _query(operation=QueryOperation.NEXT_DEADLINE),
        exact_columns=True,
        reference_date=date(2026, 7, 24),
    )
    assert report.result.value == "2026-07-30"
    assert report.result.rows[0]["Договор"] == "Договор А"


def test_due_within(contracts):
    report = answer_query(
        contracts,
        _query(operation=QueryOperation.DUE_WITHIN, horizon_days=30),
        exact_columns=True,
        reference_date=date(2026, 7, 24),
    )
    assert report.result.value == 2  # 07-30 and 08-10


def test_overdue(contracts):
    report = answer_query(
        contracts,
        _query(operation=QueryOperation.OVERDUE),
        exact_columns=True,
        reference_date=date(2026, 7, 24),
    )
    assert report.result.value == 1
    assert report.result.rows[0]["Договор"] == "Договор Г"


def test_unknown_column_cannot_answer(sales_bg):
    report = answer_query(
        sales_bg,
        _query(function=AggregateFunction.SUM, requested_metric="Печалба"),
        exact_columns=True,
    )
    assert report.result.status == ResultStatus.CANNOT_ANSWER
    assert "schema" in (report.result.message or "")


# --------------------------------------------------- duplicate headers (QA-001)


def make_duplicate_amount_workbook(path):
    """Category | Amount | Amount — the two Amount columns sum 60 and 6000."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    for col, name in enumerate(["Category", "Amount", "Amount"], start=1):
        ws.cell(row=1, column=col, value=name).font = Font(bold=True)
    rows = [("a", 10, 1000), ("b", 20, 2000), ("c", 30, 3000)]
    for i, (category, first, second) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=category)
        ws.cell(row=i, column=2, value=first)
        ws.cell(row=i, column=3, value=second)
    wb.save(path)
    return path


@pytest.fixture()
def dup_amounts(tmp_path):
    return make_duplicate_amount_workbook(tmp_path / "dup.xlsx")


def test_frame_preserves_duplicate_columns(dup_amounts):
    inventory = inventory_from_path(dup_amounts)
    table = detect_workbook_schema(inventory).tables[0]
    frame, _, _ = load_table_frame(inventory, table)
    assert list(frame.columns) == ["Category", "Amount (B)", "Amount (C)"]
    assert frame["Amount (B)"].sum() == 60
    assert frame["Amount (C)"].sum() == 6000


def test_duplicate_value_column_needs_confirmation_then_uses_choice(dup_amounts):
    query = _query(
        operation=QueryOperation.AGGREGATE,
        function=AggregateFunction.SUM,
        requested_metric="Amount",
    )
    first = answer_query(dup_amounts, query, exact_columns=True)
    assert first.result.status == ResultStatus.NEEDS_CONFIRMATION
    assert first.result.candidate_kind == "value_column"
    names = [c.column_name for c in first.result.candidates]
    assert names == ["Amount (column B)", "Amount (column C)"]

    low = answer_query(dup_amounts, query, exact_columns=True, choices=[1])
    assert low.result.value == 60
    assert low.result.provenance.value_column == "Amount (B)"

    high = answer_query(dup_amounts, query, exact_columns=True, choices=[2])
    assert high.result.value == 6000
    assert high.result.provenance.value_column == "Amount (C)"


def test_duplicate_filter_column_hits_chosen_physical_column(dup_amounts):
    query = _query(
        operation=QueryOperation.LIST_ROWS,
        filters=[
            QueryFilter(
                column="Amount", operator=FilterOperator.GREATER_THAN, value=500
            )
        ],
    )
    first = answer_query(dup_amounts, query, exact_columns=True)
    assert first.result.status == ResultStatus.NEEDS_CONFIRMATION
    assert first.result.candidate_kind == "filter_column"

    low = answer_query(dup_amounts, query, exact_columns=True, choices=[1])
    assert low.result.value == 0  # first Amount column: 10/20/30, none > 500

    high = answer_query(dup_amounts, query, exact_columns=True, choices=[2])
    assert high.result.value == 3  # second Amount column: 1000/2000/3000


def test_duplicate_group_by_column_hits_chosen_physical_column(dup_amounts):
    query = _query(
        operation=QueryOperation.AGGREGATE,
        function=AggregateFunction.SUM,
        requested_metric="Amount",
        group_by=["Amount"],
    )
    # choices: value column -> B (10/20/30), group column -> C (1000/2000/3000)
    report = answer_query(dup_amounts, query, exact_columns=True, choices=[1, 2])
    groups = {row.key["Amount (C)"]: row.value for row in report.result.groups}
    assert groups == {1000: 10, 2000: 20, 3000: 30}


def test_single_column_still_resolves_without_prompt(sales_bg_simple):
    report = answer_query(
        sales_bg_simple,
        _query(function=AggregateFunction.SUM, requested_metric="Оборот"),
        exact_columns=True,
    )
    assert report.result.status != ResultStatus.NEEDS_CONFIRMATION
    assert report.result.provenance.value_column == "Оборот"


# ------------------------------------------------ missing filter value (QA-003)


@pytest.mark.parametrize(
    "operator",
    [
        FilterOperator.GREATER_THAN,
        FilterOperator.LESS_THAN,
        FilterOperator.GREATER_OR_EQUAL,
        FilterOperator.LESS_OR_EQUAL,
        FilterOperator.EQUALS,
        FilterOperator.NOT_EQUALS,
        FilterOperator.CONTAINS,
        FilterOperator.YEAR_EQUALS,
        FilterOperator.BEFORE,
        FilterOperator.AFTER,
    ],
)
def test_filter_missing_value_raises_clean_error(sales_bg_simple, operator):
    inventory = inventory_from_path(sales_bg_simple)
    table = detect_workbook_schema(inventory).tables[0]
    column = table.column("Оборот")
    resolved = ResolvedQuery(
        operation=QueryOperation.LIST_ROWS,
        table=table,
        filters=[(column, QueryFilter(column="Оборот", operator=operator, value=None))],
    )
    with pytest.raises(ValueError, match="missing a value"):
        execute_query(inventory, resolved)


def test_filter_missing_value_is_cannot_answer_via_service(sales_bg_simple):
    report = answer_query(
        sales_bg_simple,
        _query(
            operation=QueryOperation.AGGREGATE,
            function=AggregateFunction.SUM,
            requested_metric="Оборот",
            filters=[
                QueryFilter(
                    column="Оборот", operator=FilterOperator.GREATER_THAN, value=None
                )
            ],
        ),
        exact_columns=True,
    )
    assert report.result.status == ResultStatus.CANNOT_ANSWER
    assert "missing a value" in (report.result.message or "")


def test_blank_filters_still_work_without_value(sales_bg_simple):
    report = answer_query(
        sales_bg_simple,
        _query(
            operation=QueryOperation.LIST_ROWS,
            filters=[
                QueryFilter(column="Оборот", operator=FilterOperator.IS_BLANK)
            ],
        ),
        exact_columns=True,
    )
    assert report.result.status != ResultStatus.CANNOT_ANSWER
    assert report.result.value == 2  # the two blank revenue rows
