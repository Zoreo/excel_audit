"""EA-RND-001 / EA-RND-002 rounding-drift rules (milestone 3.5, D15-D18)."""

from __future__ import annotations

import os
import subprocess
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.utils.cell import coordinate_to_tuple

import excel_auditor
from excel_auditor.analysis.dependency_graph import DependencyGraph
from excel_auditor.analysis.pattern_detection import detect_pattern_anomalies
from excel_auditor.analysis.rules import AuditContext, run_all_rules
from excel_auditor.analysis.rules.rounding import _display_decimals
from excel_auditor.analysis.workbook_inventory import inventory_from_path
from excel_auditor.models import CellRecord, Finding, SheetInventory, WorkbookInventory

FMT2 = "#,##0.00"


def _formatted_workbook(path: Path, cells: dict[str, tuple[Any, str | None]]) -> Path:
    """Real file with number formats; string values starting '=' are formulas."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Фактури"
    for coordinate, (value, fmt) in cells.items():
        ws[coordinate] = value
        if fmt is not None:
            ws[coordinate].number_format = fmt
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def _record(
    coordinate: str, value: Any, fmt: str | None = None, formula: str | None = None
) -> CellRecord:
    row, column = coordinate_to_tuple(coordinate)
    return CellRecord(
        coordinate=coordinate,
        row=row,
        column=column,
        data_type="f" if formula else "n",
        value=value,
        formula=formula,
        number_format=fmt or "General",
    )


def _inventory(
    cells: dict[str, CellRecord], *, full_precision: bool | None = None
) -> WorkbookInventory:
    sheet = SheetInventory(
        name="Фактури",
        index=0,
        max_row=max((c.row for c in cells.values()), default=0),
        max_column=max((c.column for c in cells.values()), default=0),
        cells=cells,
    )
    return WorkbookInventory(workbook_id="wb", sheets=[sheet], full_precision=full_precision)


def _run(inventory: WorkbookInventory) -> list[Finding]:
    graph = DependencyGraph.build(inventory)
    anomalies = detect_pattern_anomalies(inventory)
    return run_all_rules(
        AuditContext(inventory=inventory, graph=graph, pattern_anomalies=anomalies)
    )


def _rnd_findings(inventory: WorkbookInventory, rule_id: str = "EA-RND-001") -> list[Finding]:
    return [f for f in _run(inventory) if f.rule_id == rule_id]


# ------------------------------------------------------------ D15: decimals


def test_display_decimals_parsing():
    assert _display_decimals("#,##0.00") == 2
    assert _display_decimals("0.000") == 3
    assert _display_decimals("#,##0") == 0
    assert _display_decimals("#,##0.00 €") == 2
    assert _display_decimals('#,##0.00 "лв"') == 2
    assert _display_decimals("€#,##0.00") == 2
    assert _display_decimals("[$лв-402] #,##0.00") == 2
    assert _display_decimals("#,##0.00_);[Red](#,##0.00)") == 2  # first section only
    assert _display_decimals(None) is None
    assert _display_decimals("") is None
    assert _display_decimals("General") is None
    assert _display_decimals("@") is None  # text
    assert _display_decimals("yyyy-mm-dd") is None  # date
    assert _display_decimals("0.00%") is None  # percent is out of scope (D15)
    assert _display_decimals("0.00E+00") is None  # scientific
    assert _display_decimals("0.0#") is None  # variable decimals - not fixed


# ------------------------------------------- acceptance 1: hardcoded total


def test_hardcoded_total_drift_detected(tmp_path: Path):
    path = _formatted_workbook(
        tmp_path / "drift.xlsx",
        {
            "A1": (12.344, FMT2),
            "A2": (10.333, FMT2),
            "A3": (8.328, FMT2),
            "A4": (31.005, FMT2),
        },
    )
    findings = _rnd_findings(inventory_from_path(path))
    assert len(findings) == 1
    finding = findings[0]
    assert finding.location is not None and finding.location.coordinate == "A4"
    assert finding.severity == "medium"
    assert finding.confidence == "high"
    # Sign convention (spec): drift = displayed components − displayed total.
    assert finding.evidence["drift"] == "-0.01"
    assert finding.evidence["displayed_components_sum"] == "31.00"
    assert finding.evidence["displayed_total"] == "31.01"  # 31.005 rounds half-up
    assert "displayed components − displayed total" in finding.description
    residue = finding.evidence["residue_cells"]
    assert [r["cell"] for r in residue] == ["A1", "A2", "A3"]
    assert residue[0] == {"cell": "A1", "stored": 12.344, "displayed": "12.34"}


# ---------------------------------- acceptance 2: SUM total, cached value


def test_formula_total_with_cached_value():
    cells = {
        "A1": _record("A1", 12.344, FMT2),
        "A2": _record("A2", 10.333, FMT2),
        "A3": _record("A3", 8.328, FMT2),
        "A4": _record("A4", 31.005, FMT2, formula="=SUM(A1:A3)"),
    }
    findings = _rnd_findings(_inventory(cells))
    assert len(findings) == 1
    assert findings[0].evidence["drift"] == "-0.01"
    assert findings[0].evidence["range"] == "A1:A3"
    assert findings[0].location is not None and findings[0].location.coordinate == "A4"


# --------------------------------------------- acceptance 3: clean control


def test_exact_two_decimal_values_no_finding(tmp_path: Path):
    path = _formatted_workbook(
        tmp_path / "clean.xlsx",
        {
            "A1": (12.34, FMT2),
            "A2": (10.33, FMT2),
            "A3": (8.33, FMT2),
            "A4": (31.00, FMT2),
        },
    )
    assert _rnd_findings(inventory_from_path(path)) == []


# ------------------------------------ acceptance 4: half-up, not banker's


def test_half_away_from_zero_rounding(tmp_path: Path):
    # 2.345 displays as 2.35 (ties away from zero). Python's float round()
    # yields 2.34 here, which would flip the drift sign to -0.01.
    path = _formatted_workbook(
        tmp_path / "halfup.xlsx",
        {
            "A1": (2.345, FMT2),
            "A2": (2.345, FMT2),
            "A3": (2.31, FMT2),
            "A4": (7.0, FMT2),
        },
    )
    findings = _rnd_findings(inventory_from_path(path))
    assert len(findings) == 1
    assert findings[0].evidence["displayed_components_sum"] == "7.01"
    assert findings[0].evidence["displayed_total"] == "7.00"
    assert findings[0].evidence["drift"] == "0.01"
    assert [r["cell"] for r in findings[0].evidence["residue_cells"]] == ["A1", "A2"]


# ------------------------------------- acceptance 5: zero-decimals format


def test_zero_decimal_format_whole_unit_drift(tmp_path: Path):
    # D16 requires >= 3 populated components, so the spec's two-component
    # example (100.4 + 100.4 = 200.8) is extended to three equal components:
    # displayed 100+100+100 = 300 vs total 301.2 -> displayed 301.
    path = _formatted_workbook(
        tmp_path / "leva.xlsx",
        {
            "A1": (100.4, '#,##0 "лв"'),
            "A2": (100.4, '#,##0 "лв"'),
            "A3": (100.4, '#,##0 "лв"'),
            "A4": (301.2, '#,##0 "лв"'),
        },
    )
    findings = _rnd_findings(inventory_from_path(path))
    assert len(findings) == 1
    assert findings[0].evidence["drift"] == "-1"
    assert findings[0].evidence["displayed_components_sum"] == "300"
    assert findings[0].evidence["displayed_total"] == "301"
    assert findings[0].evidence["currency"] == "BGN"


def test_two_components_below_minimum_not_flagged(tmp_path: Path):
    # The literal criterion-5 numbers: two components stay below the D16
    # minimum of three, so no finding (conservative by design).
    path = _formatted_workbook(
        tmp_path / "leva2.xlsx",
        {"A1": (100.4, "#,##0"), "A2": (100.4, "#,##0"), "A3": (200.8, "#,##0")},
    )
    assert _rnd_findings(inventory_from_path(path)) == []


# ------------------------------- acceptance 6: unparseable component format


def test_general_component_skips_range(tmp_path: Path):
    path = _formatted_workbook(
        tmp_path / "mixed.xlsx",
        {
            "A1": (12.344, FMT2),
            "A2": (10.333, None),  # General - unparseable display decimals
            "A3": (8.328, FMT2),
            "A4": (31.005, FMT2),
        },
    )
    assert _rnd_findings(inventory_from_path(path)) == []


# ------------------------------- acceptance 7: formula with no cached value


def test_formula_total_without_cached_value(tmp_path: Path):
    # Plain openpyxl-generated file: formulas carry no cached values.
    path = _formatted_workbook(
        tmp_path / "nocache.xlsx",
        {
            "A1": (12.344, FMT2),
            "A2": (10.333, FMT2),
            "A3": (8.328, FMT2),
            "A4": ("=SUM(A1:A3)", FMT2),
        },
    )
    assert _rnd_findings(inventory_from_path(path)) == []  # skip silently, no crash


# ------------------------------------------- acceptance 8: percent formats


def test_percent_formatted_range_skipped(tmp_path: Path):
    path = _formatted_workbook(
        tmp_path / "percent.xlsx",
        {
            "A1": (0.12344, "0.00%"),
            "A2": (0.10333, "0.00%"),
            "A3": (0.08328, "0.00%"),
            "A4": (0.31005, "0.00%"),
        },
    )
    assert _rnd_findings(inventory_from_path(path)) == []


# ------------------------------------------ acceptance 9: precision setting


def test_full_precision_false_fires_workbook_finding():
    findings = _rnd_findings(_inventory({}, full_precision=False), "EA-RND-002")
    assert len(findings) == 1
    assert findings[0].location is None  # workbook-level
    assert findings[0].severity == "medium"
    assert findings[0].confidence == "high"
    assert findings[0].evidence == {"full_precision": False}


def test_full_precision_true_or_absent_stays_silent():
    assert _rnd_findings(_inventory({}, full_precision=True), "EA-RND-002") == []
    assert _rnd_findings(_inventory({}, full_precision=None), "EA-RND-002") == []


def test_inventory_captures_full_precision(tmp_path: Path):
    wb = Workbook()
    wb.active["A1"] = 1
    wb.calculation.fullPrecision = False
    path = tmp_path / "fp.xlsx"
    wb.save(path)
    assert inventory_from_path(path).full_precision is False

    plain = Workbook()
    plain.active["A1"] = 1
    plain_path = tmp_path / "plain.xlsx"
    plain.save(plain_path)
    assert inventory_from_path(plain_path).full_precision is None


# --------------------------------------------- acceptance 10: D18 cross-link


def test_cross_link_with_overwritten_formula(tmp_path: Path):
    # Row 5 holds SUM totals across columns B..F; D5 was overwritten with a
    # hardcoded total that also drifts by a cent -> EA-PAT-001 and EA-RND-001
    # both fire on D5 and cross-reference each other.
    cells: dict[str, tuple[Any, str | None]] = {}
    for col in ("B", "C", "E", "F"):
        for row in range(1, 5):
            cells[f"{col}{row}"] = (float(row), FMT2)
        cells[f"{col}5"] = (f"=SUM({col}1:{col}4)", FMT2)
    cells.update(
        {
            "D1": (12.344, FMT2),
            "D2": (10.333, FMT2),
            "D3": (8.328, FMT2),
            "D4": (10.0, FMT2),
            "D5": (41.005, FMT2),
        }
    )
    path = _formatted_workbook(tmp_path / "xlink.xlsx", cells)
    findings = _run(inventory_from_path(path))

    rnd = [f for f in findings if f.rule_id == "EA-RND-001"]
    assert len(rnd) == 1
    assert rnd[0].location is not None and rnd[0].location.coordinate == "D5"
    assert rnd[0].evidence["related_finding_rule_ids"] == ["EA-PAT-001"]

    pat = [
        f
        for f in findings
        if f.rule_id == "EA-PAT-001" and f.location and f.location.coordinate == "D5"
    ]
    assert len(pat) == 1
    assert pat[0].evidence["related_finding_rule_ids"] == ["EA-RND-001"]
    assert pat[0].evidence["related_finding_note"] == (
        "likely manual rounding adjustment — see EA-RND-001"
    )


# ------------------------------------------------ acceptance 11: determinism

_DRIVER = """\
import sys
from datetime import UTC, datetime
from pathlib import Path

from excel_auditor.reporting.html_report import render_audit_html
from excel_auditor.reporting.json_report import to_json
from excel_auditor.services import audit_workbook

report = audit_workbook(sys.argv[2], generated_at=datetime(2026, 7, 24, 12, 0, 0, tzinfo=UTC))
out_dir = Path(sys.argv[1])
(out_dir / "report.json").write_bytes(to_json(report).encode("utf-8"))
(out_dir / "report.html").write_bytes(render_audit_html(report).encode("utf-8"))
"""


def test_report_with_rnd_findings_byte_identical_across_hash_seeds(tmp_path: Path):
    fixture = _formatted_workbook(
        tmp_path / "drift.xlsx",
        {
            "A1": (12.344, FMT2),
            "A2": (10.333, FMT2),
            "A3": (8.328, FMT2),
            "A4": (31.005, FMT2),
        },
    )
    driver = tmp_path / "driver.py"
    driver.write_text(_DRIVER, encoding="utf-8")
    src = str(Path(excel_auditor.__file__).resolve().parents[1])

    outputs = []
    for seed in (0, 42):
        out_dir = tmp_path / f"seed{seed}"
        out_dir.mkdir()
        env = dict(os.environ)
        env["PYTHONHASHSEED"] = str(seed)
        env["PYTHONPATH"] = src + os.pathsep + env.get("PYTHONPATH", "")
        subprocess.run(
            [sys.executable, str(driver), str(out_dir), str(fixture)], check=True, env=env
        )
        outputs.append(
            ((out_dir / "report.json").read_bytes(), (out_dir / "report.html").read_bytes())
        )

    assert b"EA-RND-001" in outputs[0][0]  # the finding is actually in the report
    assert outputs[0][0] == outputs[1][0]
    assert outputs[0][1] == outputs[1][1]
