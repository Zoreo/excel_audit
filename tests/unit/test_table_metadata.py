"""T12: exact Excel Table metadata (headerRowCount/totalsRowCount, D14).

Fixtures are built programmatically with openpyxl: a real Excel `Table`
object whose XML declares the header/totals split, so the schema and the
dependency graph can use exact metadata instead of heuristics.
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.table import Table

from excel_auditor.analysis.dependency_graph import DependencyGraph
from excel_auditor.analysis.query import load_table_frame
from excel_auditor.analysis.schema import detect_workbook_schema
from excel_auditor.analysis.workbook_inventory import inventory_from_path
from excel_auditor.models.workbook import TableInfo

_DATA = [
    ("Widget", "EU", 100),
    ("Gadget", "US", 200),
    ("Doohickey", "EU", 300),
]


def _make_table_workbook(
    path: Path,
    *,
    totals_row_count: int | None = None,
    totals_value: object = "=SUBTOTAL(109,SalesTbl[Amount])",
    formula_cell: str | None = "=SUM(SalesTbl[Amount])",
) -> Path:
    """Sales table in A1:C4 (+ totals row 5 when totals_row_count) as a real
    Excel Table named SalesTbl; optional structured-ref formula in E1."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    for col, header in zip("ABC", ["Item", "Region", "Amount"], strict=True):
        ws[f"{col}1"] = header
    for i, (item, region, amount) in enumerate(_DATA, start=2):
        ws[f"A{i}"] = item
        ws[f"B{i}"] = region
        ws[f"C{i}"] = amount
    last_row = 1 + len(_DATA)
    if totals_row_count:
        last_row += 1
        ws[f"A{last_row}"] = "Total"
        ws[f"C{last_row}"] = totals_value
    table = Table(displayName="SalesTbl", ref=f"A1:C{last_row}")
    if totals_row_count is not None:
        table.totalsRowCount = totals_row_count
    ws.add_table(table)
    if formula_cell is not None:
        ws["E1"] = formula_cell
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


# ------------------------------------------------------------- inventory


def test_inventory_captures_exact_table_metadata(tmp_path: Path):
    path = _make_table_workbook(tmp_path / "totals.xlsx", totals_row_count=1)
    inventory = inventory_from_path(path)
    (info,) = inventory.sheets[0].tables
    assert info.name == "SalesTbl"
    assert info.ref == "A1:C5"
    assert info.header_row_count == 1
    assert info.totals_row_count == 1


def test_inventory_coerces_missing_totals_to_zero(tmp_path: Path):
    # openpyxl leaves totalsRowCount=None unless set; must land as 0.
    path = _make_table_workbook(tmp_path / "no_totals.xlsx", totals_row_count=None)
    inventory = inventory_from_path(path)
    (info,) = inventory.sheets[0].tables
    assert info.ref == "A1:C4"
    assert info.header_row_count == 1
    assert info.totals_row_count == 0


def test_table_info_defaults_mirror_excel():
    info = TableInfo(name="T", ref="A1:B2", sheet_name="S")
    assert info.header_row_count == 1
    assert info.totals_row_count == 0


# ----------------------------------------------------- schema (criterion 1)


def test_schema_exact_split_with_totals_row(tmp_path: Path):
    path = _make_table_workbook(tmp_path / "totals.xlsx", totals_row_count=1)
    schema = detect_workbook_schema(inventory_from_path(path))
    (table,) = schema.tables
    assert table.ref == "A1:C5"
    assert table.header_rows == [1]
    assert table.data_start_row == 2
    assert table.data_end_row == 5
    assert table.total_rows == [5]  # exact, from the Table XML
    assert table.row_count == 3  # totals row excluded
    assert any("exact (from Excel Table metadata)" in note for note in table.notes)
    # Columns come from the Table's own ref, not the surrounding block.
    assert [c.name for c in table.columns] == ["Item", "Region", "Amount"]


def test_query_sum_excludes_totals_row(tmp_path: Path):
    # Literal grand total in the totals row: including it would double the sum.
    path = _make_table_workbook(
        tmp_path / "literal_totals.xlsx",
        totals_row_count=1,
        totals_value=600,
        formula_cell=None,
    )
    inventory = inventory_from_path(path)
    (table,) = detect_workbook_schema(inventory).tables
    frame, rows_total, totals_excluded = load_table_frame(inventory, table)
    assert totals_excluded == 1
    assert list(frame["Item"]) == ["Widget", "Gadget", "Doohickey"]  # no "Total" row
    assert frame["Amount"].sum() == 600  # not 1200


# ------------------------------------------- dependency graph (criterion 2)


def test_structured_ref_edges_stop_before_totals(tmp_path: Path):
    path = _make_table_workbook(tmp_path / "totals.xlsx", totals_row_count=1)
    inventory = inventory_from_path(path)
    graph = DependencyGraph.build(inventory)
    # =SUM(SalesTbl[Amount]) covers the data rows only: C2:C4, never C5.
    assert graph.precedents[("Sales", "E1")] == {
        ("Sales", "C2"),
        ("Sales", "C3"),
        ("Sales", "C4"),
    }
    # The totals row's own SUBTOTAL doesn't include itself: no self-loop.
    assert graph.precedents[("Sales", "C5")] == {
        ("Sales", "C2"),
        ("Sales", "C3"),
        ("Sales", "C4"),
    }
    assert not graph.self_loops
    assert not graph.unresolved_name_cells


# --------------------------------------------- totalsRowCount=None (crit. 3)


def test_totals_none_behaves_as_zero(tmp_path: Path):
    path = _make_table_workbook(tmp_path / "no_totals.xlsx", totals_row_count=None)
    inventory = inventory_from_path(path)

    schema = detect_workbook_schema(inventory)
    (table,) = schema.tables
    assert table.total_rows == []
    assert table.data_start_row == 2
    assert table.data_end_row == 4
    assert table.row_count == 3  # all data rows kept
    assert any("exact (from Excel Table metadata)" in note for note in table.notes)

    graph = DependencyGraph.build(inventory)
    assert graph.precedents[("Sales", "E1")] == {
        ("Sales", "C2"),
        ("Sales", "C3"),
        ("Sales", "C4"),
    }
    assert not graph.unresolved_name_cells


# ------------------------------------------- non-Table blocks (criterion 4)


def test_plain_block_keeps_heuristics(tmp_path: Path):
    # Same grid, including a "Total" row, but WITHOUT an Excel Table object:
    # the heuristic path must handle it exactly as before.
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    for col, header in zip("ABC", ["Item", "Region", "Amount"], strict=True):
        ws[f"{col}1"] = header
    for i, (item, region, amount) in enumerate(_DATA, start=2):
        ws[f"A{i}"] = item
        ws[f"B{i}"] = region
        ws[f"C{i}"] = amount
    ws["A5"] = "Total"
    ws["C5"] = 600
    path = tmp_path / "plain.xlsx"
    wb.save(path)

    schema = detect_workbook_schema(inventory_from_path(path))
    (table,) = schema.tables
    assert table.header_rows == [1]
    assert table.total_rows == [5]  # heuristic keyword detection, as today
    assert table.row_count == 3
    assert not any("exact" in note for note in table.notes)
    assert any("look like totals" in note for note in table.notes)


# --------------------------------------------------- headerRowCount edge


def test_header_row_count_zero(tmp_path: Path):
    # headerRowCount=0 never round-trips through openpyxl defaults, so patch
    # the inventory: schema falls back to positional column names and the
    # graph refuses to guess a named column without an in-sheet header row.
    path = _make_table_workbook(tmp_path / "no_totals.xlsx", totals_row_count=None)
    inventory = inventory_from_path(path)
    inventory.sheets[0].tables[0].header_row_count = 0

    schema = detect_workbook_schema(inventory)
    (table,) = schema.tables
    assert table.header_rows == []
    assert table.data_start_row == 1
    assert table.row_count == 4  # the former header row is data now
    assert [c.name for c in table.columns] == ["Column A", "Column B", "Column C"]

    graph = DependencyGraph.build(inventory)
    # SalesTbl[Amount] cannot be matched without a header row: unknown marker.
    assert ("Sales", "E1") in graph.unresolved_name_cells
    assert graph.precedents[("Sales", "E1")] == set()
