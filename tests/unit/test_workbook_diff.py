from pathlib import Path

from openpyxl import Workbook

from conftest import make_workbook
from excel_auditor.analysis.workbook_diff import compare_inventories
from excel_auditor.analysis.workbook_inventory import inventory_from_path
from excel_auditor.models import ChangeType, StructuralChangeType


def _changes_by_key(cell_changes):
    return {(c.sheet_name, c.coordinate): c for c in cell_changes}


def test_demo_diff_detects_planted_changes(old_inventory, new_inventory):
    structural, cell_changes = compare_inventories(old_inventory, new_inventory)
    changes = _changes_by_key(cell_changes)

    # 1: formula overwritten with hardcoded number
    d7 = changes[("Revenue Forecast", "D7")]
    assert d7.change_type == ChangeType.FORMULA_TO_CONSTANT
    assert d7.new_value == 2600

    # 2: wrong-row copy is a formula change
    assert changes[("Cash Flow", "B9")].change_type == ChangeType.FORMULA_CHANGED

    # 3: shrunk total range
    d14 = changes[("Revenue Forecast", "D14")]
    assert d14.change_type == ChangeType.FORMULA_CHANGED
    assert d14.old_formula == "=SUM(D2:D13)"
    assert d14.new_formula == "=SUM(D2:D12)"

    # 6: changed assumption value
    b3 = changes[("Assumptions", "B3")]
    assert b3.change_type == ChangeType.VALUE_CHANGED
    assert b3.old_value == 0.05 and b3.new_value == 0.07

    # 7: formatting-only change
    assert changes[("P&L", "B6")].change_type == ChangeType.FORMATTING_ONLY

    # 10: downstream formula change
    assert changes[("P&L", "B3")].change_type == ChangeType.FORMULA_CHANGED

    # 4: hidden sheet added
    added = [
        c
        for c in structural
        if c.change_type == StructuralChangeType.SHEET_ADDED and c.sheet_name == "Adjustments"
    ]
    assert len(added) == 1
    assert added[0].details["visibility"] == "hidden"

    # 9: volatile formula appears as an added formula
    assert changes[("Summary", "B8")].change_type == ChangeType.FORMULA_ADDED


def test_unchanged_cells_not_reported(old_inventory, new_inventory):
    _, cell_changes = compare_inventories(old_inventory, new_inventory)
    keys = {(c.sheet_name, c.coordinate) for c in cell_changes}
    assert ("Revenue Forecast", "D3") not in keys
    assert ("P&L", "B5") not in keys


def test_sheet_rename_inferred(tmp_path: Path):
    cells = {"A1": "x", "B2": 5, "C3": "=B2*2"}
    old = make_workbook(tmp_path / "old.xlsx", {"Data": dict(cells)})
    new = make_workbook(tmp_path / "new.xlsx", {"Data2024": dict(cells)})
    structural, cell_changes = compare_inventories(
        inventory_from_path(old, workbook_id="old"),
        inventory_from_path(new, workbook_id="new"),
    )
    renames = [c for c in structural if c.change_type == StructuralChangeType.SHEET_RENAMED]
    assert len(renames) == 1
    assert renames[0].details == {"old_name": "Data", "new_name": "Data2024"}
    assert cell_changes == []  # renamed sheet content is identical


def test_sheet_reorder_detected(tmp_path: Path):
    old = make_workbook(tmp_path / "old.xlsx", {"One": {"A1": 1}, "Two": {"A1": 2}})
    new = make_workbook(tmp_path / "new.xlsx", {"Two": {"A1": 2}, "One": {"A1": 1}})
    structural, _ = compare_inventories(
        inventory_from_path(old, workbook_id="old"),
        inventory_from_path(new, workbook_id="new"),
    )
    assert any(c.change_type == StructuralChangeType.SHEETS_REORDERED for c in structural)


def _diff(old_path: Path, new_path: Path):
    return compare_inventories(
        inventory_from_path(old_path, workbook_id="old"),
        inventory_from_path(new_path, workbook_id="new"),
    )


def _renames(structural):
    return [c for c in structural if c.change_type == StructuralChangeType.SHEET_RENAMED]


def _removed_added(structural):
    removed = [c.sheet_name for c in structural
               if c.change_type == StructuralChangeType.SHEET_REMOVED]
    added = [c.sheet_name for c in structural
             if c.change_type == StructuralChangeType.SHEET_ADDED]
    return removed, added


# Five formula cells: enough signal for inferred rename matching (D3).
_FORMULA_SHEET = {
    "A1": 10,
    "A2": 20,
    "B1": "=A1*2",
    "B2": "=A2*2",
    "B3": "=B1+B2",
    "B4": "=SUM(A1:A2)",
    "B5": "=B3-B4",
    "C1": 100,
}


def test_inferred_rename_with_value_edit_reports_both(tmp_path: Path):
    """EXCEL-001: rename + edit must not erase the sheet's changes."""
    old = make_workbook(tmp_path / "old.xlsx", {"Data": dict(_FORMULA_SHEET)})
    new_cells = dict(_FORMULA_SHEET)
    new_cells["C1"] = 999  # single value edit -> content no longer identical
    new = make_workbook(tmp_path / "new.xlsx", {"Data2024": new_cells})

    structural, cell_changes = _diff(old, new)
    renames = _renames(structural)
    assert len(renames) == 1
    assert renames[0].details == {
        "old_name": "Data", "new_name": "Data2024", "inferred": True
    }
    assert "content similarity" in renames[0].description
    removed, added = _removed_added(structural)
    assert removed == [] and added == []
    changes = _changes_by_key(cell_changes)
    assert changes[("Data2024", "C1")].change_type == ChangeType.VALUE_CHANGED
    assert changes[("Data2024", "C1")].new_value == 999


def test_inferred_rename_with_formula_edit_at_similarity_boundary(tmp_path: Path):
    """4 of 5 formulas shared = 0.80 similarity: still paired, edit reported."""
    old = make_workbook(tmp_path / "old.xlsx", {"Data": dict(_FORMULA_SHEET)})
    new_cells = dict(_FORMULA_SHEET)
    new_cells["B5"] = "=B3+B4"  # single formula edit
    new = make_workbook(tmp_path / "new.xlsx", {"Data2024": new_cells})

    structural, cell_changes = _diff(old, new)
    renames = _renames(structural)
    assert len(renames) == 1
    assert renames[0].details["inferred"] is True
    changes = _changes_by_key(cell_changes)
    assert changes[("Data2024", "B5")].change_type == ChangeType.FORMULA_CHANGED
    assert changes[("Data2024", "B5")].old_formula == "=B3-B4"
    assert changes[("Data2024", "B5")].new_formula == "=B3+B4"


def test_genuinely_different_sheets_stay_removed_added(tmp_path: Path):
    """Q1 -> Q2 with different content must not be paired as a rename."""
    q1 = {f"B{i}": f"=A{i}*{i}" for i in range(1, 7)}
    q2 = {f"B{i}": f"=A{i}+{i + 10}" for i in range(1, 7)}
    old = make_workbook(tmp_path / "old.xlsx", {"Q1": q1})
    new = make_workbook(tmp_path / "new.xlsx", {"Q2": q2})

    structural, _ = _diff(old, new)
    assert _renames(structural) == []
    removed, added = _removed_added(structural)
    assert removed == ["Q1"] and added == ["Q2"]


def test_ambiguous_candidates_pair_nothing(tmp_path: Path):
    """Two removed and two added sheets, all equally similar: margin fails."""
    def sheet(marker: int) -> dict:
        cells = dict(_FORMULA_SHEET)
        cells["C1"] = marker  # distinct values, identical formulas
        return cells

    old = make_workbook(tmp_path / "old.xlsx", {"R1": sheet(1), "R2": sheet(2)})
    new = make_workbook(tmp_path / "new.xlsx", {"N1": sheet(3), "N2": sheet(4)})

    structural, _ = _diff(old, new)
    assert _renames(structural) == []
    removed, added = _removed_added(structural)
    assert sorted(removed) == ["R1", "R2"]
    assert sorted(added) == ["N1", "N2"]


def test_two_clear_renames_both_pair(tmp_path: Path):
    """Two removed + two added, each with a unique best match: both pair."""
    sheet_a = dict(_FORMULA_SHEET)
    sheet_b = {f"B{i}": f"=A{i}+{i}" for i in range(1, 7)} | {"A1": 1}
    old = make_workbook(tmp_path / "old.xlsx", {"R1": dict(sheet_a), "R2": dict(sheet_b)})
    new_a = dict(sheet_a)
    new_a["C1"] = 999
    new_b = dict(sheet_b)
    new_b["A1"] = 2
    new = make_workbook(tmp_path / "new.xlsx", {"N1": new_a, "N2": new_b})

    structural, _ = _diff(old, new)
    renames = _renames(structural)
    pairs = {(r.details["old_name"], r.details["new_name"]) for r in renames}
    assert pairs == {("R1", "N1"), ("R2", "N2")}
    assert all(r.details["inferred"] is True for r in renames)
    removed, added = _removed_added(structural)
    assert removed == [] and added == []


def test_few_formula_cells_never_pair_inferred(tmp_path: Path):
    """Fewer than 5 formula cells: not enough signal for inference."""
    cells = {"A1": 1, "B1": "=A1*2", "B2": "=A1*3", "B3": "=A1*4", "B4": "=A1*5"}
    old = make_workbook(tmp_path / "old.xlsx", {"Data": dict(cells)})
    new_cells = dict(cells)
    new_cells["A1"] = 2  # edit so the exact-signature path cannot match
    new = make_workbook(tmp_path / "new.xlsx", {"Data2024": new_cells})

    structural, _ = _diff(old, new)
    assert _renames(structural) == []
    removed, added = _removed_added(structural)
    assert removed == ["Data"] and added == ["Data2024"]


def test_pure_rename_is_reference_change_free(tmp_path: Path):
    """P2-1: cross-sheet references to a renamed sheet are not formula changes."""
    inputs = {"A1": 1, "A2": 2, "B1": "=A1+A2"}
    raw = {"A1": 5}
    old = make_workbook(
        tmp_path / "old.xlsx",
        {
            "Inputs": dict(inputs),
            "Raw Data": dict(raw),
            "Summary": {
                "A1": "=Inputs!B1",
                "A2": "=SUM(Inputs!A1:A2)",
                "A3": "='Raw Data'!A1",
            },
        },
    )
    new = make_workbook(
        tmp_path / "new.xlsx",
        {
            "Assumptions": dict(inputs),
            "Raw Data 2024": dict(raw),
            "Summary": {
                "A1": "=Assumptions!B1",
                "A2": "=SUM(Assumptions!A1:A2)",
                "A3": "='Raw Data 2024'!A1",
            },
        },
    )

    structural, cell_changes = _diff(old, new)
    pairs = {(r.details["old_name"], r.details["new_name"]) for r in _renames(structural)}
    assert pairs == {("Inputs", "Assumptions"), ("Raw Data", "Raw Data 2024")}
    assert cell_changes == []  # a pure rename produces zero cell changes


def test_rename_plus_real_formula_edit_still_reported(tmp_path: Path):
    """P2-1: only the pure rename substitution is neutralized."""
    inputs = {"A1": 1, "A2": 2, "B1": "=A1+A2"}
    old = make_workbook(
        tmp_path / "old.xlsx",
        {"Inputs": dict(inputs), "Summary": {"A1": "=Inputs!B1", "A2": "=Inputs!A1"}},
    )
    new = make_workbook(
        tmp_path / "new.xlsx",
        {
            "Assumptions": dict(inputs),
            "Summary": {"A1": "=Assumptions!B1*2", "A2": "=Assumptions!A1"},
        },
    )

    structural, cell_changes = _diff(old, new)
    assert len(_renames(structural)) == 1
    changes = _changes_by_key(cell_changes)
    assert set(changes) == {("Summary", "A1")}  # A2 is the pure rename, silent
    assert changes[("Summary", "A1")].change_type == ChangeType.FORMULA_CHANGED


def test_structural_changes_emit_in_sorted_sheet_order(tmp_path: Path):
    """EXCEL-004: matched-sheet iteration must not depend on set ordering."""

    def build(path: Path, *, changed: bool) -> Path:
        wb = Workbook()
        default = wb.active
        for index in range(1, 8):
            name = f"S{index}"
            ws = default if index == 1 else wb.create_sheet()
            ws.title = name
            ws["A1"] = index
            if changed:
                if name in ("S2", "S5"):
                    ws.sheet_state = "hidden"
                if name == "S3":
                    ws.merge_cells("A2:B3")
                if name == "S6":
                    ws.merge_cells("C1:D2")
                if name in ("S4", "S7"):
                    ws.row_dimensions[3].hidden = True
        wb.save(path)
        return path

    old = build(tmp_path / "old.xlsx", changed=False)
    new = build(tmp_path / "new.xlsx", changed=True)
    structural, _ = compare_inventories(
        inventory_from_path(old, workbook_id="old"),
        inventory_from_path(new, workbook_id="new"),
    )
    assert [(c.change_type, c.sheet_name) for c in structural] == [
        (StructuralChangeType.SHEET_VISIBILITY_CHANGED, "S2"),
        (StructuralChangeType.MERGED_RANGES_CHANGED, "S3"),
        (StructuralChangeType.HIDDEN_ROWS_CHANGED, "S4"),
        (StructuralChangeType.SHEET_VISIBILITY_CHANGED, "S5"),
        (StructuralChangeType.MERGED_RANGES_CHANGED, "S6"),
        (StructuralChangeType.HIDDEN_ROWS_CHANGED, "S7"),
    ]


def test_constant_to_formula(tmp_path: Path):
    old = make_workbook(tmp_path / "old.xlsx", {"S": {"A1": 10, "A2": 20}})
    new = make_workbook(tmp_path / "new.xlsx", {"S": {"A1": 10, "A2": "=A1*2"}})
    _, cell_changes = compare_inventories(
        inventory_from_path(old, workbook_id="old"),
        inventory_from_path(new, workbook_id="new"),
    )
    changes = _changes_by_key(cell_changes)
    assert changes[("S", "A2")].change_type == ChangeType.CONSTANT_TO_FORMULA
