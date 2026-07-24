from excel_auditor.analysis.resolution import (
    concepts_in_text,
    concepts_of,
    normalize,
    resolve_date_column,
    resolve_exact_column,
    resolve_metric,
    transliterate,
)
from excel_auditor.analysis.schema import detect_workbook_schema
from excel_auditor.analysis.workbook_inventory import inventory_from_path


def _tables(path):
    return detect_workbook_schema(inventory_from_path(path)).tables


def test_normalize():
    assert normalize("  Нетен   Оборот! ") == "нетен оборот"
    assert normalize("Net-Revenue (EUR)") == "net revenue eur"


def test_transliterate():
    assert transliterate("оборот") == "oborot"
    assert transliterate("щастие") == "shtastie"


def test_concepts_of_bulgarian_and_english():
    assert "revenue" in concepts_of("Оборот")
    assert "revenue" in concepts_of("Revenue")
    assert "revenue" in concepts_of("Oborot")  # transliterated header
    assert "net_revenue" in concepts_of("Нетен оборот")
    # the net column also mentions the generic term
    assert "revenue" in concepts_of("Нетен оборот")
    assert "customer" in concepts_of("Клиент")


def test_concepts_in_text_prefers_specific():
    concepts = concepts_in_text("Какъв е общият нетен оборот за 2025?")
    assert concepts[0] == "net_revenue"
    assert "revenue" not in concepts  # subsumed by the longer alias


def test_exact_resolution(sales_bg):
    tables = _tables(sales_bg)
    resolution = resolve_exact_column("Оборот", tables)
    assert resolution.status == "resolved"
    assert resolution.single is not None
    assert resolution.single.column.name == "Оборот"


def test_metric_resolution_is_ambiguous_with_net_column(sales_bg):
    tables = _tables(sales_bg)
    resolution = resolve_metric("оборот", tables)
    assert resolution.status == "ambiguous"
    names = {m.column.name for m in resolution.matches}
    assert names == {"Оборот", "Нетен оборот"}


def test_metric_resolution_specific_net(sales_bg):
    tables = _tables(sales_bg)
    resolution = resolve_metric("нетен оборот", tables)
    assert resolution.status == "resolved"
    assert resolution.single.column.name == "Нетен оборот"


def test_metric_resolution_single_column(sales_bg_simple):
    tables = _tables(sales_bg_simple)
    resolution = resolve_metric("revenue", tables)
    assert resolution.status == "resolved"
    assert resolution.single.column.name == "Оборот"


def test_translit_header_resolution(tmp_path):
    from conftest import make_sales_workbook

    path = make_sales_workbook(
        tmp_path / "translit.xlsx", headers="translit", net_column=False
    )
    tables = _tables(path)
    resolution = resolve_metric("оборот", tables)
    assert resolution.status == "resolved"
    assert resolution.single.column.name == "Oborot"


def test_date_resolution(sales_bg):
    tables = _tables(sales_bg)
    resolution = resolve_date_column(tables)
    assert resolution.status == "resolved"
    assert resolution.single.column.name == "Дата"


def test_date_resolution_ambiguous_with_two_dates(tmp_path):
    from conftest import make_sales_workbook

    path = make_sales_workbook(tmp_path / "two_dates.xlsx", second_date=True)
    tables = _tables(path)
    resolution = resolve_date_column(tables)
    assert resolution.status == "ambiguous"
    assert len(resolution.matches) == 2


def make_duplicate_amount_workbook(path):
    """Category | Amount | Amount — the two Amount columns sum 60 and 6000."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    for col, name in enumerate(["Category", "Amount", "Amount"], start=1):
        ws.cell(row=1, column=col, value=name).font = Font(bold=True)
    rows = [("a", 10, 1000), ("b", 20, 2000), ("c", 30, 3000)]
    for i, (category, first, second) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=category)
        ws.cell(row=i, column=2, value=first)
        ws.cell(row=i, column=3, value=second)
    wb.save(path)
    return path


def test_duplicate_headers_exact_resolution_is_ambiguous(tmp_path):
    tables = _tables(make_duplicate_amount_workbook(tmp_path / "dup.xlsx"))
    resolution = resolve_exact_column("Amount", tables)
    assert resolution.status == "ambiguous"
    assert [m.column.letter for m in resolution.matches] == ["B", "C"]
    assert {m.column.name for m in resolution.matches} == {"Amount"}


def test_duplicate_headers_metric_resolution_is_ambiguous(tmp_path):
    tables = _tables(make_duplicate_amount_workbook(tmp_path / "dup.xlsx"))
    resolution = resolve_metric("amount", tables)
    assert resolution.status == "ambiguous"
    assert [m.column.letter for m in resolution.matches] == ["B", "C"]


def test_forecast_and_actual_are_ambiguous(tmp_path):
    from conftest import make_sales_workbook

    path = make_sales_workbook(
        tmp_path / "forecast.xlsx", net_column=False, forecast_columns=True
    )
    tables = _tables(path)
    resolution = resolve_metric("revenue", tables)
    assert resolution.status == "ambiguous"
    names = {m.column.name for m in resolution.matches}
    assert {"Revenue Actual", "Revenue Forecast"} <= names
