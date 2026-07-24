import pytest

from excel_auditor.llm import UnsupportedQuestionError, get_parser
from excel_auditor.llm.mock_parser import MockIntentParser
from excel_auditor.llm.rule_parser import RuleBasedIntentParser
from excel_auditor.models.query import (
    AggregateFunction,
    FilterOperator,
    QueryAction,
    QueryOperation,
    ResultStatus,
    SpreadsheetQuery,
)
from excel_auditor.query_service import answer_query, inspect_schema


@pytest.fixture(scope="module")
def parser() -> RuleBasedIntentParser:
    return RuleBasedIntentParser()


@pytest.fixture(scope="module")
def bg_schema(sales_bg):
    return inspect_schema(sales_bg).workbook_schema


# ------------------------------------------------------------- rule parsing


def test_parse_total_revenue_en(parser, bg_schema):
    query = parser.parse("What was total revenue in 2025?", bg_schema)
    assert query.operation == QueryOperation.AGGREGATE
    assert query.function == AggregateFunction.SUM
    assert query.requested_metric == "revenue"
    assert query.filters[0].operator == FilterOperator.YEAR_EQUALS
    assert query.filters[0].value == 2025


def test_parse_total_revenue_bg(parser, bg_schema):
    query = parser.parse("Какъв е общият оборот за 2025?", bg_schema)
    assert query.function == AggregateFunction.SUM
    assert query.requested_metric == "revenue"
    assert query.filters[0].value == 2025


def test_parse_specific_net_revenue(parser, bg_schema):
    query = parser.parse("Какъв е общият нетен оборот за 2025?", bg_schema)
    assert query.requested_metric == "net_revenue"


def test_parse_unique_customers(parser, bg_schema):
    query = parser.parse("How many unique customers are there?", bg_schema)
    assert query.function == AggregateFunction.DISTINCT_COUNT
    assert query.requested_metric == "customer"


def test_parse_group_by(parser, bg_schema):
    query = parser.parse("Show revenue by region", bg_schema)
    assert query.function == AggregateFunction.SUM
    assert query.group_by == ["region"]


def test_parse_period_comparison(parser, bg_schema):
    query = parser.parse("Compare revenue in 2024 and 2025", bg_schema)
    assert query.operation == QueryOperation.COMPARE_PERIODS
    assert query.period_comparison.period_a == 2024
    assert query.period_comparison.period_b == 2025


def test_parse_deadlines(parser, bg_schema):
    assert (
        parser.parse("What is the next contract deadline?", bg_schema).operation
        == QueryOperation.NEXT_DEADLINE
    )
    query = parser.parse("Колко договора изтичат до 30 дни?", bg_schema)
    assert query.operation == QueryOperation.DUE_WITHIN
    assert query.horizon_days == 30
    assert (
        parser.parse("Which invoices are overdue?", bg_schema).operation
        == QueryOperation.OVERDUE
    )


def test_parse_inspection_and_audit(parser, bg_schema):
    assert (
        parser.parse("What sheets are in this workbook?", bg_schema).action
        == QueryAction.INSPECT_WORKBOOK
    )
    assert (
        parser.parse("Audit this workbook.", bg_schema).action
        == QueryAction.AUDIT_WORKBOOK
    )
    trace = parser.parse("Trace dependencies from Sales!D7", bg_schema)
    assert trace.action == QueryAction.TRACE_DEPENDENCIES
    assert trace.cell_reference == "Sales!D7"


def test_parse_rejects_open_ended(parser, bg_schema):
    for question in (
        "Tell me everything interesting about this company.",
        "Explain why the business is failing.",
        "Find fraud.",
    ):
        with pytest.raises(UnsupportedQuestionError):
            parser.parse(question, bg_schema)


# --------------------------------------------------------------- end to end


def test_ask_bg_requires_confirmation_then_answers(sales_bg):
    parser = RuleBasedIntentParser()
    schema = inspect_schema(sales_bg).workbook_schema
    query = parser.parse("Какъв е общият оборот за 2025?", schema)

    first = answer_query(sales_bg, query, exact_columns=False)
    assert first.result.status == ResultStatus.NEEDS_CONFIRMATION
    assert first.result.candidate_kind == "value_column"
    names = [c.column_name for c in first.result.candidates]
    assert names == ["Оборот", "Нетен оборот"]

    confirmed = answer_query(sales_bg, query, exact_columns=False, choices=[1])
    assert confirmed.result.value == 628400

    net = answer_query(sales_bg, query, exact_columns=False, choices=[2])
    assert net.result.value == 565560  # net 2025 sum


def test_ask_simple_workbook_no_confirmation(sales_bg_simple):
    parser = RuleBasedIntentParser()
    schema = inspect_schema(sales_bg_simple).workbook_schema
    query = parser.parse("What was total revenue in 2025?", schema)
    report = answer_query(sales_bg_simple, query, exact_columns=False)
    assert report.result.status != ResultStatus.NEEDS_CONFIRMATION
    assert report.result.value == 628400


def test_ask_duplicate_headers_requires_confirmation_then_uses_choice(tmp_path):
    """QA-001 end to end: Category|Amount|Amount must never resolve silently."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    path = tmp_path / "dup.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    for col, name in enumerate(["Category", "Amount", "Amount"], start=1):
        ws.cell(row=1, column=col, value=name).font = Font(bold=True)
    rows = [("a", 10, 1000), ("b", 20, 2000), ("c", 30, 3000)]
    for i, (category, first, second) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=category)
        ws.cell(row=i, column=2, value=first)
        ws.cell(row=i, column=3, value=second)
    wb.save(path)

    parser = RuleBasedIntentParser()
    schema = inspect_schema(path).workbook_schema
    query = parser.parse("What is the total amount?", schema)

    first = answer_query(path, query, exact_columns=False)
    assert first.result.status == ResultStatus.NEEDS_CONFIRMATION
    assert first.result.candidate_kind == "value_column"
    names = [c.column_name for c in first.result.candidates]
    assert names == ["Amount (column B)", "Amount (column C)"]

    low = answer_query(path, query, exact_columns=False, choices=[1])
    assert low.result.value == 60

    high = answer_query(path, query, exact_columns=False, choices=[2])
    assert high.result.value == 6000


def test_ask_inspect_workbook(sales_bg):
    parser = RuleBasedIntentParser()
    schema = inspect_schema(sales_bg).workbook_schema
    query = parser.parse("What sheets are in this workbook?", schema)
    report = answer_query(sales_bg, query, exact_columns=False)
    assert report.result.status == ResultStatus.VERIFIED
    assert "Sales" in (report.result.message or "")


def test_mock_parser_used_for_tests(sales_bg):
    mock = MockIntentParser()
    canned = SpreadsheetQuery(
        function=AggregateFunction.SUM, requested_metric="Оборот"
    )
    mock.register("canned?", canned)
    schema = inspect_schema(sales_bg).workbook_schema
    assert mock.parse("canned?", schema) is canned
    assert mock.seen == ["canned?"]
    with pytest.raises(UnsupportedQuestionError):
        mock.parse("unknown", schema)


def test_get_parser_selection(monkeypatch):
    monkeypatch.setenv("EXCEL_AUDITOR_INTENT_PARSER", "mock")
    assert isinstance(get_parser(), MockIntentParser)
    monkeypatch.delenv("EXCEL_AUDITOR_INTENT_PARSER")
    assert isinstance(get_parser(), RuleBasedIntentParser)
    assert isinstance(get_parser("mock"), MockIntentParser)
