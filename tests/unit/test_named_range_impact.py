"""EXCEL-003: defined names and structured table references resolve to real
graph edges; only genuinely unresolvable tokens keep the unknown-impact marker."""

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table

from conftest import make_workbook
from excel_auditor.analysis.dependency_graph import DependencyGraph, impact_for
from excel_auditor.analysis.severity import classify_cell_change
from excel_auditor.analysis.workbook_inventory import inventory_from_path
from excel_auditor.models import ChangeType, Confidence, DependencyImpact, Severity


def _build_workbook(
    path: Path,
    sheets: dict[str, dict[str, Any]],
    *,
    names: tuple[tuple[str, str], ...] = (),
    sheet_names: tuple[tuple[str, str, str], ...] = (),
    tables: tuple[tuple[str, str, str], ...] = (),
) -> Path:
    """Workbook builder with defined names and tables.

    names: (name, refers_to) workbook-scoped; sheet_names: (sheet, name,
    refers_to) sheet-scoped; tables: (sheet, display_name, ref).
    """
    wb = Workbook()
    default = wb.active
    worksheets = {}
    for index, (title, cells) in enumerate(sheets.items()):
        ws = default if index == 0 else wb.create_sheet()
        ws.title = title
        for coordinate, value in cells.items():
            ws[coordinate] = value
        worksheets[title] = ws
    for name, refers_to in names:
        wb.defined_names.add(DefinedName(name, attr_text=refers_to))
    for sheet, name, refers_to in sheet_names:
        worksheets[sheet].defined_names.add(DefinedName(name, attr_text=refers_to))
    for sheet, display_name, ref in tables:
        worksheets[sheet].add_table(Table(displayName=display_name, ref=ref))
    wb.save(path)
    return path


def _graph(path: Path, **kwargs):
    inventory = inventory_from_path(path)
    return inventory, DependencyGraph.build(inventory, **kwargs)


# ------------------------------------------------------------ defined names


def test_named_range_parity_with_direct_reference(tmp_path: Path):
    # Twelve dependents so the >= 10 severity escalation is exercised.
    named_cells = {"A1": 100, **{f"B{r}": "=MyInput+1" for r in range(1, 13)}}
    direct_cells = {"A1": 100, **{f"B{r}": "=A1+1" for r in range(1, 13)}}
    named_inv, named_graph = _graph(
        _build_workbook(tmp_path / "named.xlsx", {"Model": named_cells},
                        names=(("MyInput", "Model!$A$1"),))
    )
    direct_inv, direct_graph = _graph(
        _build_workbook(tmp_path / "direct.xlsx", {"Model": direct_cells})
    )

    named_impact = impact_for(named_graph, named_inv, ("Model", "A1"))
    direct_impact = impact_for(direct_graph, direct_inv, ("Model", "A1"))
    assert named_impact.direct_dependent_count == direct_impact.direct_dependent_count == 12
    assert named_impact.transitive_dependent_count == direct_impact.transitive_dependent_count
    assert not named_impact.has_unresolved_names

    named_outcome = classify_cell_change(ChangeType.VALUE_CHANGED, impact=named_impact)
    direct_outcome = classify_cell_change(ChangeType.VALUE_CHANGED, impact=direct_impact)
    assert named_outcome == direct_outcome == (Severity.MEDIUM, Confidence.HIGH)


def test_multi_area_name_resolves_all_areas(tmp_path: Path):
    _, graph = _graph(
        _build_workbook(
            tmp_path / "multi.xlsx",
            {"Model": {"A1": 1, "A2": 2, "A3": 3, "C1": 4, "E1": "=SUM(MultiArea)"}},
            names=(("MultiArea", "Model!$A$1:$A$3,Model!$C$1"),),
        )
    )
    assert graph.precedents[("Model", "E1")] == {
        ("Model", "A1"), ("Model", "A2"), ("Model", "A3"), ("Model", "C1"),
    }
    assert not graph.unresolved_name_cells


def test_constant_valued_names_no_edges_no_marker(tmp_path: Path):
    _, graph = _graph(
        _build_workbook(
            tmp_path / "const.xlsx",
            {"Model": {"B1": "=ConstStr*2", "B2": "=ConstNum+1"}},
            names=(("ConstStr", '"0.05"'), ("ConstNum", "0.05")),
        )
    )
    assert not graph.unresolved_name_cells  # fully understood: nothing understated
    assert ("Model", "B1") not in graph.precedents
    assert ("Model", "B2") not in graph.precedents


def test_sheet_scoped_name_shadows_global(tmp_path: Path):
    _, graph = _graph(
        _build_workbook(
            tmp_path / "scoped.xlsx",
            {
                "Model": {"A1": 100, "D1": "=Rate*2"},
                "Local": {"B1": 7, "C1": "=Rate*2"},
            },
            names=(("Rate", "Model!$A$1"),),
            sheet_names=(("Local", "Rate", "Local!$B$1"),),
        )
    )
    # On 'Local' the sheet-scoped Rate wins over the same-named global ...
    assert graph.precedents[("Local", "C1")] == {("Local", "B1")}
    # ... while other sheets still see the workbook-scoped definition.
    assert graph.precedents[("Model", "D1")] == {("Model", "A1")}
    assert not graph.unresolved_name_cells


def test_ref_error_name_keeps_marker(tmp_path: Path):
    inventory, graph = _graph(
        _build_workbook(
            tmp_path / "broken.xlsx",
            {"Model": {"B1": "=Broken+1"}},
            names=(("Broken", "Model!#REF!"),),
        )
    )
    assert graph.unresolved_name_cells == {("Model", "B1")}
    assert ("Model", "B1") not in graph.precedents  # no mis-resolved edge
    assert impact_for(graph, inventory, ("Model", "B1")).has_unresolved_names


def test_unknown_name_keeps_marker_and_capped_confidence(tmp_path: Path):
    inventory, graph = _graph(
        _build_workbook(tmp_path / "unknown.xlsx", {"Model": {"A1": 1, "B1": "=NoSuchName*2"}})
    )
    assert graph.unresolved_name_cells == {("Model", "B1")}
    impact = impact_for(graph, inventory, ("Model", "A1"))
    assert impact.has_unresolved_names
    severity, confidence = classify_cell_change(ChangeType.VALUE_CHANGED, impact=impact)
    assert severity == Severity.LOW  # no aggressive escalation ...
    assert confidence == Confidence.MEDIUM  # ... but no confident "low impact"


def test_name_routed_self_reference_is_circular(tmp_path: Path):
    inventory, graph = _graph(
        _build_workbook(
            tmp_path / "namecycle.xlsx",
            {"Model": {"A1": "=MyName+1"}},
            names=(("MyName", "Model!$A$1"),),
        )
    )
    assert graph.self_loops == {("Model", "A1")}
    assert [("Model", "A1")] in graph.cycles()
    assert impact_for(graph, inventory, ("Model", "A1")).is_circular


def test_name_range_respects_max_range_cells(tmp_path: Path):
    cells: dict[str, Any] = {f"A{r}": r for r in range(1, 51)}
    cells["C1"] = "=SUM(BigRange)"
    _, graph = _graph(
        _build_workbook(
            tmp_path / "big.xlsx", {"Model": cells},
            names=(("BigRange", "Model!$A$1:$A$50"),),
        ),
        max_range_cells=10,
    )
    assert graph.truncated_ranges == 1
    assert len(graph.precedents[("Model", "C1")]) <= 10


# -------------------------------------------------------- structured tables


def test_table_column_resolves_to_data_subrange(tmp_path: Path):
    path = _build_workbook(
        tmp_path / "table.xlsx",
        {
            "Data": {
                "A1": "Item", "B1": "Amount",
                "A2": "x", "B2": 10,
                "A3": "y", "B3": 20,
                "A4": "z", "B4": 30,
                "D1": "=SUM(Table1[Amount])",
            }
        },
        tables=(("Data", "Table1", "A1:B4"),),
    )
    _, graph = _graph(path)
    # Data rows of the Amount column only: no header, no Item column.
    assert graph.precedents[("Data", "D1")] == {
        ("Data", "B2"), ("Data", "B3"), ("Data", "B4"),
    }
    assert ("Data", "D1") in graph.dependents[("Data", "B2")]
    assert not graph.unresolved_name_cells


def test_table_totals_row_excluded_and_no_false_self_loop(tmp_path: Path):
    path = _build_workbook(
        tmp_path / "totals.xlsx",
        {
            "Data": {
                "A1": "Item", "B1": "Amount",
                "A2": "x", "B2": 10,
                "A3": "y", "B3": 20,
                "A4": "Total", "B4": "=SUBTOTAL(109,Table1[Amount])",
                "D1": "=SUM(Table1[Amount])",
            }
        },
        tables=(("Data", "Table1", "A1:B4"),),
    )
    _, graph = _graph(path)
    # The totals row (self-referential aggregate) is not part of the data.
    assert graph.precedents[("Data", "D1")] == {("Data", "B2"), ("Data", "B3")}
    assert graph.precedents[("Data", "B4")] == {("Data", "B2"), ("Data", "B3")}
    assert not graph.self_loops
    assert not graph.unresolved_name_cells


def test_unsupported_table_specifier_keeps_marker(tmp_path: Path):
    path = _build_workbook(
        tmp_path / "spec.xlsx",
        {
            "Data": {
                "A1": "Item", "B1": "Amount",
                "A2": "x", "B2": 10,
                "D1": "=SUM(Table1[#All])",
                "D2": "=SUM(Table1[NoSuchColumn])",
            }
        },
        tables=(("Data", "Table1", "A1:B2"),),
    )
    _, graph = _graph(path)
    assert graph.unresolved_name_cells == {("Data", "D1"), ("Data", "D2")}
    assert ("Data", "D1") not in graph.precedents
    assert ("Data", "D2") not in graph.precedents


def test_undefined_table_reference_keeps_marker(tmp_path: Path):
    path = make_workbook(
        tmp_path / "notable.xlsx",
        {"Data": {"A1": 1, "A2": 2, "B1": "=SUM(Table1[Amount])"}},
    )
    inventory, graph = _graph(path)
    assert graph.unresolved_name_cells == {("Data", "B1")}
    assert impact_for(graph, inventory, ("Data", "A1")).has_unresolved_names


# ------------------------------------------------------------ marker hygiene


def test_no_defined_names_no_marker(tmp_path: Path):
    path = make_workbook(
        tmp_path / "plain.xlsx",
        {"Model": {"A1": 100, "B1": "=A1*2", "C1": "=B1+1"}},
    )
    inventory, graph = _graph(path)
    assert not graph.unresolved_name_cells

    impact = impact_for(graph, inventory, ("Model", "A1"))
    assert not impact.has_unresolved_names
    assert impact.transitive_dependent_count == 2

    severity, confidence = classify_cell_change(ChangeType.VALUE_CHANGED, impact=impact)
    assert severity == Severity.LOW
    assert confidence == Confidence.HIGH  # no false alarm


def test_demonstrated_impact_still_escalates_with_full_confidence():
    # When the escalation already fired, the classification is not understated:
    # severity goes up as before and confidence stays high.
    impact = DependencyImpact(
        transitive_dependent_count=25, has_unresolved_names=True
    )
    severity, confidence = classify_cell_change(ChangeType.VALUE_CHANGED, impact=impact)
    assert severity == Severity.MEDIUM
    assert confidence == Confidence.HIGH
