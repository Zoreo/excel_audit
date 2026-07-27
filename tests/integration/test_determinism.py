"""EXCEL-004: report generation must be deterministic.

Golden byte-comparison of JSON and HTML output across subprocesses with
different PYTHONHASHSEED values, content-hash workbook ids, and the
``generated_at`` injection point (service parameter + CLI passthrough).
"""

from __future__ import annotations

import hashlib
import json
import os
import subprocess
import sys
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook

import excel_auditor
from excel_auditor.cli import main
from excel_auditor.services import audit_workbook, compare_workbooks

_SRC = str(Path(excel_auditor.__file__).resolve().parents[1])

_STAMP = datetime(2026, 7, 24, 12, 0, 0, tzinfo=UTC)

_DRIVER = """\
import sys
from datetime import UTC, datetime
from pathlib import Path

from excel_auditor.reporting.html_report import render_audit_html, render_comparison_html
from excel_auditor.reporting.json_report import to_json
from excel_auditor.services import audit_workbook, compare_workbooks

stamp = datetime(2026, 7, 24, 12, 0, 0, tzinfo=UTC)
mode, out_dir = sys.argv[1], Path(sys.argv[2])
if mode == "audit":
    report = audit_workbook(sys.argv[3], generated_at=stamp)
    html = render_audit_html(report)
else:
    report = compare_workbooks(sys.argv[3], sys.argv[4], generated_at=stamp)
    html = render_comparison_html(report)
(out_dir / "report.json").write_bytes(to_json(report).encode("utf-8"))
(out_dir / "report.html").write_bytes(html.encode("utf-8"))
"""


def _make_workbook(
    path: Path,
    sheets: dict[str, dict[str, Any]],
    *,
    hidden_sheets: tuple[str, ...] = (),
    merged: dict[str, list[str]] | None = None,
    hidden_rows: dict[str, list[int]] | None = None,
    hidden_columns: dict[str, list[str]] | None = None,
) -> Path:
    """Small workbook builder; strings starting with '=' become formulas."""
    wb = Workbook()
    default = wb.active
    for index, (name, cells) in enumerate(sheets.items()):
        ws = default if index == 0 else wb.create_sheet()
        ws.title = name
        for coordinate, value in cells.items():
            ws[coordinate] = value
        if name in hidden_sheets:
            ws.sheet_state = "hidden"
        for ref in (merged or {}).get(name, []):
            ws.merge_cells(ref)
        for row in (hidden_rows or {}).get(name, []):
            ws.row_dimensions[row].hidden = True
        for column in (hidden_columns or {}).get(name, []):
            ws.column_dimensions[column].hidden = True
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def _base_sheets() -> dict[str, dict[str, Any]]:
    return {f"S{i}": {"A1": i, "A2": i * 10, "B1": f"=A1*{i}+A2"} for i in range(1, 8)}


def _run_in_subprocess(tmp_path: Path, seed: int, mode: str, *paths: Path) -> tuple[bytes, bytes]:
    out_dir = tmp_path / f"{mode}-seed{seed}"
    out_dir.mkdir()
    driver = tmp_path / "driver.py"
    driver.write_text(_DRIVER, encoding="utf-8")
    env = dict(os.environ)
    env["PYTHONHASHSEED"] = str(seed)
    env["PYTHONPATH"] = _SRC + os.pathsep + env.get("PYTHONPATH", "")
    subprocess.run(
        [sys.executable, str(driver), mode, str(out_dir), *map(str, paths)],
        check=True,
        env=env,
    )
    return (out_dir / "report.json").read_bytes(), (out_dir / "report.html").read_bytes()


# ------------------------------------------------- golden byte comparisons


def test_audit_byte_identical_across_hash_seeds(tmp_path: Path):
    fixture = _make_workbook(tmp_path / "fixture.xlsx", _base_sheets())
    runs = [_run_in_subprocess(tmp_path, seed, "audit", fixture) for seed in (0, 42)]
    baseline_json, baseline_html = runs[0]
    for run_json, run_html in runs[1:]:
        assert run_json == baseline_json
        assert run_html == baseline_html


def _row_table(units: list[int]) -> dict[str, Any]:
    """A sheet above the D7 row-alignment gate (>= 5 data rows + total)."""
    cells: dict[str, Any] = {}
    for offset, value in enumerate(units):
        row = offset + 2
        cells[f"A{row}"] = value
        cells[f"B{row}"] = f"=A{row}*2"
    total_row = len(units) + 2
    cells[f"A{total_row}"] = "Total"
    cells[f"B{total_row}"] = f"=SUM(B2:B{total_row - 1})"
    return cells


def test_compare_byte_identical_across_hash_seeds(tmp_path: Path):
    # Several matched sheets with visibility/merged/hidden changes so that
    # structural-change ordering is actually exercised, plus a row-aligned
    # sheet with an inserted row so the schema-v3 ROWS_INSERTED path (D7/D8
    # signature alignment) is covered by the byte comparison too.
    old_sheets = _base_sheets()
    old_sheets["Rows"] = _row_table([10, 20, 30, 40, 50, 60])
    old = _make_workbook(tmp_path / "old.xlsx", old_sheets)
    new_sheets = _base_sheets()
    new_sheets["S1"]["B1"] = "=A1*99"
    new_sheets["Rows"] = _row_table([10, 20, 30, 99, 40, 50, 60])  # row inserted
    new = _make_workbook(
        tmp_path / "new.xlsx",
        new_sheets,
        hidden_sheets=("S2", "S5"),
        merged={"S3": ["A3:B4"], "S6": ["C1:D2"]},
        hidden_rows={"S4": [3], "S7": [2]},
        hidden_columns={"S5": ["D"]},
    )
    runs = [_run_in_subprocess(tmp_path, seed, "compare", old, new) for seed in (0, 1, 42)]
    baseline_json, baseline_html = runs[0]
    for run_json, run_html in runs[1:]:
        assert run_json == baseline_json
        assert run_html == baseline_html


# --------------------------------------------------- content-hash workbook id


def test_workbook_id_is_stable_content_hash(tmp_path: Path):
    fixture = _make_workbook(tmp_path / "same.xlsx", {"Data": {"A1": 1, "B1": "=A1*2"}})
    expected = hashlib.sha256(fixture.read_bytes()).hexdigest()
    first = audit_workbook(fixture)
    second = audit_workbook(fixture)
    assert first.workbook.workbook_id == expected
    assert second.workbook.workbook_id == expected


def test_workbook_id_changes_with_content(tmp_path: Path):
    original = _make_workbook(tmp_path / "wb.xlsx", {"Data": {"A1": 1, "B1": "=A1*2"}})
    original_id = audit_workbook(original).workbook.workbook_id
    _make_workbook(tmp_path / "wb.xlsx", {"Data": {"A1": 2, "B1": "=A1*2"}})
    changed_id = audit_workbook(original).workbook.workbook_id
    assert changed_id != original_id


def test_compare_keeps_old_new_workbook_ids(tmp_path: Path):
    old = _make_workbook(tmp_path / "old.xlsx", {"Data": {"A1": 1}})
    new = _make_workbook(tmp_path / "new.xlsx", {"Data": {"A1": 2}})
    comparison = compare_workbooks(old, new)
    assert comparison.old_workbook.workbook_id == "old"
    assert comparison.new_workbook.workbook_id == "new"


# --------------------------------------------------------- CLI passthrough


def test_cli_audit_generated_at_passthrough(tmp_path: Path):
    fixture = _make_workbook(tmp_path / "fixture.xlsx", _base_sheets())
    out_json = tmp_path / "audit.json"
    code = main(
        [
            "audit", str(fixture),
            "--generated-at", "2026-07-24T12:00:00+00:00",
            "--output-dir", str(tmp_path / "artifacts"),
            "--json-output", str(out_json),
        ]
    )
    assert code == 0
    payload = json.loads(out_json.read_text(encoding="utf-8"))
    assert datetime.fromisoformat(payload["generated_at"]) == _STAMP


def test_cli_compare_generated_at_passthrough(tmp_path: Path):
    old = _make_workbook(tmp_path / "old.xlsx", {"Data": {"A1": 1}})
    new = _make_workbook(tmp_path / "new.xlsx", {"Data": {"A1": 2}})
    out_json = tmp_path / "cmp.json"
    code = main(
        [
            "compare", str(old), str(new),
            "--generated-at", "2026-07-24T12:00:00+00:00",
            "--output-dir", str(tmp_path / "artifacts"),
            "--json-output", str(out_json),
        ]
    )
    assert code == 0
    payload = json.loads(out_json.read_text(encoding="utf-8"))
    assert datetime.fromisoformat(payload["generated_at"]) == _STAMP
