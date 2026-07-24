"""Cross-fix interaction tests (lead-owned, per the remediation plan).

Each test exercises two independently fixed findings together to prove the
fixes compose: a wrong answer prevented by one fix must stay prevented when
the code path of another fix is also in play.
"""

import json
import os
import subprocess
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.styles import Font

import excel_auditor
from excel_auditor.analysis.rules import base as rules_base
from excel_auditor.llm.rule_parser import RuleBasedIntentParser
from excel_auditor.models.query import FilterOperator, ResultStatus
from excel_auditor.query_service import answer_query, inspect_schema
from excel_auditor.services import audit_workbook

_SRC = str(Path(excel_auditor.__file__).resolve().parents[1])


def _duplicate_amount_workbook(path: Path) -> Path:
    """Category|Amount|Amount with divergent data in the two Amount columns."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    for col, name in enumerate(["Category", "Amount", "Amount"], start=1):
        ws.cell(row=1, column=col, value=name).font = Font(bold=True)
    rows = [("a", 100, 1000), ("b", 200, 2000), ("c", 700, 7000), ("d", 900, 9000)]
    for i, (category, first, second) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=category)
        ws.cell(row=i, column=2, value=first)
        ws.cell(row=i, column=3, value=second)
    wb.save(path)
    return path


def test_threshold_query_on_duplicate_headers_confirms_then_filters_chosen_column(
    tmp_path,
):
    """QA-001 x QA-002: a threshold question against duplicate headers must
    first surface the ambiguity, then filter the CHOSEN physical column."""
    path = _duplicate_amount_workbook(tmp_path / "dup_threshold.xlsx")

    parser = RuleBasedIntentParser()
    schema = inspect_schema(path).workbook_schema
    query = parser.parse("show rows where amount over 500", schema)

    # QA-002: the threshold filter must be built, not silently dropped.
    assert [(f.column, f.operator, f.value) for f in query.filters] == [
        ("amount", FilterOperator.GREATER_THAN, 500.0)
    ]

    # QA-001: two physical Amount columns -> never resolved silently.
    first = answer_query(path, query, exact_columns=False)
    assert first.result.status == ResultStatus.NEEDS_CONFIRMATION
    labels = [c.column_name for c in first.result.candidates]
    assert labels == ["Amount (column B)", "Amount (column C)"]

    # Column B: 100/200/700/900 -> two rows exceed 500.
    chose_b = answer_query(path, query, exact_columns=False, choices=[1])
    assert chose_b.result.status == ResultStatus.VERIFIED
    assert chose_b.result.provenance.rows_included == 2
    assert any("500" in f for f in chose_b.result.provenance.filters)

    # Column C: 1000/2000/7000/9000 -> every row exceeds 500.
    chose_c = answer_query(path, query, exact_columns=False, choices=[2])
    assert chose_c.result.status == ResultStatus.VERIFIED
    assert chose_c.result.provenance.rows_included == 4


class _ExplodingRule(rules_base.Rule):
    rule_id = "EA-TST-999"
    title = "Always crashes"
    description = "Test-only rule that raises unconditionally."

    def run(self, ctx):  # noqa: ARG002 - signature fixed by the Rule ABC
        raise RuntimeError("boom")


def test_self_loop_finding_survives_a_crashed_rule(
    tmp_path, monkeypatch: pytest.MonkeyPatch
):
    """EXCEL-002 x EXCEL-005: one rule crashing must not suppress the
    self-loop finding, and the coverage warning must coexist with it."""
    monkeypatch.setattr(rules_base, "ALL_RULES", [*rules_base.ALL_RULES, _ExplodingRule])

    wb = Workbook()
    ws = wb.active
    ws.title = "Loop"
    ws["A1"] = "=A1+1"
    path = tmp_path / "selfloop.xlsx"
    wb.save(path)

    report = audit_workbook(path)
    circular = [f for f in report.findings if f.rule_id == "EA-CIR-001"]
    assert len(circular) == 1
    assert report.failed_rules == ["EA-TST-999"]
    assert any("failed to run" in driver for driver in report.risk_drivers)


_RENAME_DRIVER = """\
import sys
from datetime import UTC, datetime
from pathlib import Path

from excel_auditor.reporting.html_report import render_comparison_html
from excel_auditor.reporting.json_report import to_json
from excel_auditor.services import compare_workbooks

stamp = datetime(2026, 7, 24, 12, 0, 0, tzinfo=UTC)
out_dir = Path(sys.argv[1])
report = compare_workbooks(sys.argv[2], sys.argv[3], generated_at=stamp)
(out_dir / "report.json").write_bytes(to_json(report).encode("utf-8"))
(out_dir / "report.html").write_bytes(render_comparison_html(report).encode("utf-8"))
"""


def _rename_fixture(path: Path, *, sheet_name: str, edited: bool) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in range(1, 8):
        ws[f"A{row}"] = row
        ws[f"B{row}"] = f"=A{row}*2"
    if edited:
        ws["A1"] = 99
    summary = wb.create_sheet()
    summary.title = "Summary"
    summary["A1"] = f"={sheet_name}!B1+1"
    wb.save(path)
    return path


def test_inferred_rename_comparison_is_deterministic(tmp_path):
    """EXCEL-001 x EXCEL-004: a rename+edit comparison must report both the
    inferred rename and the edit, byte-identically across hash seeds."""
    old = _rename_fixture(tmp_path / "old.xlsx", sheet_name="Inputs", edited=False)
    new = _rename_fixture(tmp_path / "new.xlsx", sheet_name="Assumptions", edited=True)

    outputs: list[tuple[bytes, bytes]] = []
    for seed in (0, 1, 42):
        out_dir = tmp_path / f"seed{seed}"
        out_dir.mkdir()
        driver = tmp_path / "driver.py"
        driver.write_text(_RENAME_DRIVER, encoding="utf-8")
        env = dict(os.environ)
        env["PYTHONHASHSEED"] = str(seed)
        env["PYTHONPATH"] = _SRC + os.pathsep + env.get("PYTHONPATH", "")
        subprocess.run(
            [sys.executable, str(driver), str(out_dir), str(old), str(new)],
            check=True,
            env=env,
        )
        outputs.append(
            ((out_dir / "report.json").read_bytes(), (out_dir / "report.html").read_bytes())
        )

    assert outputs[0] == outputs[1] == outputs[2]

    payload = json.loads(outputs[0][0])
    renames = [
        c
        for c in payload["structural_changes"]
        if c["change_type"] == "sheet_renamed"
    ]
    assert len(renames) == 1
    assert renames[0]["details"] == {
        "old_name": "Inputs",
        "new_name": "Assumptions",
        "inferred": True,
    }
    edits = [
        c
        for c in payload["cell_changes"]
        if c["change_type"] == "value_changed" and c["coordinate"] == "A1"
    ]
    assert len(edits) == 1
    # P2-1: Summary!A1's =Inputs!B1+1 vs =Assumptions!B1+1 is a pure rename
    # reference, not a formula change.
    assert not any(
        c["sheet_name"] == "Summary" for c in payload["cell_changes"]
    )
